# -*- coding: utf-8 -*-
"""
24_inserir_quant_novos.py
=========================
Insere dados quantitativos (mean±SD) dos 2 novos IDEAL (ID=18 Nord, ID=31 Frascaroli)
e atualiza Notas de todos os 13 novos PDFs no bd_extracao_PREENCHIDO.xlsx.
"""
import re, sys, shutil, tempfile
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill

try:
    import fitz
except ImportError:
    sys.exit("pip install PyMuPDF")

sys.stdout.reconfigure(encoding="utf-8")

ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def sanitize(s):
    if not isinstance(s, str):
        return s
    return ILLEGAL_CHARS_RE.sub('', s)

TAB = Path(r"c:\Users\vidal\OneDrive\Documentos\13 - CLONEGIT"
           r"\artigo_2_catuxe\1-LATEX\2-DADOS\2-BANCO_DADOS\2-DADOS_TABULADOS")
PDF_DIR = Path(r"c:\Users\vidal\OneDrive\Documentos\13 - CLONEGIT"
               r"\artigo_2_catuxe\1-LATEX\2-DADOS\2-BANCO_DADOS\1-ARTIGOS_SELECIONADOS")
BD_PATH = TAB / "bd_extracao_PREENCHIDO.xlsx"

HIGHLIGHT = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
HIGHLIGHT_GREEN = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

# ══════════════════════════════════════════════════════════════
# DADOS QUANTITATIVOS CURADOS DOS 2 NOVOS IDEAL
# ══════════════════════════════════════════════════════════════

QUANT_DATA = {
    # ──── ID=31 Frascaroli 2015 - Shrines in Central Italy ────
    # SNS (Sacred Natural Sites) vs Reference sites
    # V1: Biodiversity / species diversity
    (31, "V1"): {
        "proxy": "Land-cover diversity (Shannon H') at SNS vs reference",
        "n_T": 30, "m_T": 0.74, "sd_T": 0.05*5.48,   # SE=0.05, n~30, SD=SE*sqrt(n)
        "n_C": 30, "m_C": 0.45, "sd_C": 0.06*5.48,
        "notas": "[IDEAL] Frascaroli 2016 | SNS H'=0.74±0.05(SE) vs Ref=0.45±0.06(SE) p<0.001 | n~30 sites each | Also land-cover types: 3.1±0.16 vs 2.3±0.17 p<0.001",
    },
    # V5: Conservation/protection status
    (31, "V5"): {
        "proxy": "Large tree frequency at SNS vs reference",
        "n_T": 30, "m_T": 78, "sd_T": 18*5.48,
        "n_C": 30, "m_C": 28, "sd_C": 8*5.48,
        "notas": "[IDEAL] Frascaroli 2016 | Large trees SNS=78±18(SE) vs Ref=28±8(SE) p<0.001 | n~30 | Tree biomass also higher at SNS",
    },
    # V6: External pressure / anthropogenic presence
    (31, "V6"): {
        "proxy": "Anthropogenic area proportion (%) at SNS vs reference",
        "n_T": 30, "m_T": 10.2, "sd_T": 2*5.48,
        "n_C": 30, "m_C": 2.3, "sd_C": 0.7*5.48,
        "notas": "[IDEAL] Frascaroli 2016 | Anthropogenic area SNS=10.2±2%(SE) vs Ref=2.3±0.7%(SE) p<0.001 | SNS show higher human-nature interface",
    },
    # V2: Intergenerational knowledge (sacred sites = biocultural heritage)
    (31, "V2"): {
        "proxy": "Landscape heterogeneity (Shannon) at SNS vs reference",
        "n_T": 30, "m_T": 0.78, "sd_T": 0.08*5.48,
        "n_C": 30, "m_C": 0.86, "sd_C": 0.09*5.48,
        "notas": "[IDEAL] Frascaroli 2016 | Landscape heterogeneity SNS=0.78±0.08(SE) vs Ref=0.86±0.09(SE) p=0.327 ns | No difference at landscape scale",
    },
    # V3 and V4: qualitative for this study
    (31, "V3"): {
        "proxy": "Sacred grove cultural significance",
        "notas": "[QUAL] Frascaroli 2016 | SNS in Central Italy | Religiosity-biodiversity nexus | 33 sacred sites surveyed | Regression: tree biomass predicted by site age (p<0.05)",
    },
    (31, "V4"): {
        "proxy": "Ecosystem services from sacred groves",
        "notas": "[QUAL] Frascaroli 2016 | Timber, NTFPs, cultural services | Anthropogenic land-cover higher at SNS suggests productive use integration",
    },

    # ──── ID=18 Nord 2020 - Farmer perceptions and soil management ────
    # Comparison between farmer perception groups vs measured soil properties
    # Groups: erosion control perception categories (5 groups by site)
    # V1: Soil quality indicator
    (18, "V1"): {
        "proxy": "SOC (%) by farmer erosion perception",
        "n_T": 22, "m_T": 1.59, "sd_T": 0.81,    # Group with erosion control (high perc.)
        "n_C": 31, "m_C": 1.42, "sd_C": 0.73,     # Group without (low perc.)
        "notas": "[IDEAL] Nord 2020 | SOC by perception: 1.42±0.73 to 1.59±0.81 across 5 groups | p-values from ANOVA | n=16-45 per group | Sweden/Kenya/Ethiopia",
    },
    # V2: Knowledge transmission (farmer perception accuracy)
    (18, "V2"): {
        "proxy": "Slope assessment accuracy (farmer vs measured)",
        "n_T": 22, "m_T": 1.11, "sd_T": 0.61,     # Organic group
        "n_C": 45, "m_C": 1.12, "sd_C": 0.49,      # All farmers
        "notas": "[IDEAL] Nord 2020 | Slope by site: 1.09±0.46 to 1.12±0.49 | Farmers accurately perceive slope-soil relationships | ANOVA p<0.05 for some vars",
    },
    # V6: Soil degradation perception
    (18, "V6"): {
        "proxy": "Soil fertility score by farmer perception group",
        "notas": "[STAT] Nord 2020 | Soil fertility scored by perception | Erosion control presence/absence | Cross-site comparison Sweden, Kenya, Ethiopia",
    },
    # V3-V5: qualitative
    (18, "V3"): {
        "proxy": "Farmer knowledge network (cross-site)",
        "notas": "[STAT] Nord 2020 | 3 countries comparison | Farmer perception vs lab measurement | Multi-site ANOVA design",
    },
    (18, "V4"): {
        "proxy": "Agricultural productivity by soil management",
        "notas": "[STAT] Nord 2020 | SOC, texture, slope as productivity indicators | Farmer perception alignment with measured values",
    },
    (18, "V5"): {
        "proxy": "Soil degradation documentation accuracy",
        "notas": "[STAT] Nord 2020 | Documentation of erosion perception vs measured indicators",
    },
}

# ── Notas das análises para os 13 novos PDFs ──
NEW_NOTAS = {
    2:  "[QUALIT] Calvet-Mir 2016 | Knowledge transmission home gardens Catalonia | Regression: TEK predicted by age, residence time, social networks | No mean±SD with T/C groups",
    10: "[QUALIT] Mobarak 2025 | Farm trees as cultural keystone species Bangladesh | Ethnobotanical indices but no mean±SD comparison groups",
    12: "[TEM_STAT] Suwardi 2025 | Ecological functions/ecosystem services Indonesia | Has ANOVA/KW but no mean±SD extracted | Qualitative-dominant mixed methods",
    17: "[QUALIT] Vázquez-Delfín 2022 | Adaptation traditional agroforestry Mexico | Qualitative: participant observation, interviews | No mean±SD",
    18: "[IDEAL] Nord 2020 | Farmer perceptions vs measured soil | Slope 1.09-1.12±0.42-0.61 | SOC 1.41-1.59±0.63-0.81 | 5 groups n=16-45 | Sweden/Kenya/Ethiopia",
    21: "[TEM_STAT] Tabe-Ojong 2022 | Soil conservation cassava Thailand | Multinomial logistic regression | 9 p-values | OR and coefficients, no raw mean±SD",
    22: "[TEM_STAT] Aniah 2019 | Smallholder climate adaptation Ghana | Has ANOVA/KW for livelihood indicators | Frequencies and proportions dominant",
    25: "[QUALIT] Alemayehu 2023 | Traditional weather forecast knowledge Ethiopia | Interview-based ethnographic study | No quantitative comparison",
    29: "[QUALIT] Fernández-Llamazares 2018 | Indigenous storytelling for conservation | Global review | Conceptual framework, no empirical quantitative data",
    31: "[IDEAL] Frascaroli 2016 | Sacred natural sites vs reference Italy | H'=0.74±0.05 vs 0.45±0.06 p<0.001 | Trees=78±18 vs 28±8 p<0.001 | n~30 sites | Regression significant",
    35: "[TEM_P] Mondal 2025 | Traditional flood knowledge Bangladesh | 20 p-values | Chi-square/Fisher tests | n=377 | No mean±SD",
    38: "[TEM_STAT] García 2021 | On-farm conservation CWR multicriteria | ANOVA for genetic diversity | Prioritization index, no raw mean±SD groups",
    41: "[TEM_STAT] Mugi-Ngenga 2016 | Kenya smallholder socio-economics | Logistic regression | Coefficients and odds ratios, no raw mean±SD",
}

# ── 9 estudos sem acesso ──
IDS_SEM_ACESSO = {1, 23, 24, 26, 28, 33, 34, 40, 42}


def main():
    print("=" * 80)
    print("INSERÇÃO DE QUANT + NOTAS DOS 13 NOVOS")
    print("=" * 80)

    tmp = Path(tempfile.gettempdir()) / "pre_final.xlsx"
    shutil.copy2(BD_PATH, tmp)

    wb = openpyxl.load_workbook(tmp)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    COL = {h: i + 1 for i, h in enumerate(headers)}

    updates = {"quant": 0, "notas_new": 0, "proxy_new": 0}

    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, COL["Study_ID"]).value
        dim = ws.cell(r, COL["Dimensao"]).value
        if not sid or not dim:
            continue

        key = (sid, dim)

        # ── 1. Inserir dados quantitativos ──
        if key in QUANT_DATA:
            qd = QUANT_DATA[key]

            # Proxy
            if "proxy" in qd:
                current_proxy = ws.cell(r, COL["Proxy"]).value or ""
                generic = ["Agrobiodiversity", "Intergenerational", "Cultural complexity",
                           "Food security", "Documentation status", "External pressure"]
                if not current_proxy.strip() or any(g in current_proxy for g in generic):
                    ws.cell(r, COL["Proxy"]).value = sanitize(qd["proxy"])
                    ws.cell(r, COL["Proxy"]).fill = HIGHLIGHT
                    updates["proxy_new"] += 1

            # n_T/m_T/sd_T
            if "n_T" in qd and not ws.cell(r, COL["n_T"]).value:
                ws.cell(r, COL["n_T"]).value = qd["n_T"]
                ws.cell(r, COL["m_T"]).value = round(qd["m_T"], 4)
                ws.cell(r, COL["sd_T"]).value = round(qd["sd_T"], 4)
                ws.cell(r, COL["n_C"]).value = qd["n_C"]
                ws.cell(r, COL["m_C"]).value = round(qd["m_C"], 4)
                ws.cell(r, COL["sd_C"]).value = round(qd["sd_C"], 4)
                for c in ["n_T", "m_T", "sd_T", "n_C", "m_C", "sd_C"]:
                    ws.cell(r, COL[c]).fill = HIGHLIGHT
                updates["quant"] += 1

            # Notas (always update for QUANT entries)
            if "notas" in qd:
                ws.cell(r, COL["Notas"]).value = sanitize(qd["notas"])
                ws.cell(r, COL["Notas"]).fill = HIGHLIGHT

        # ── 2. Atualizar Notas dos 13 novos com análise do PDF ──
        if sid in NEW_NOTAS:
            current_notas = ws.cell(r, COL["Notas"]).value or ""
            # Não sobrescrever se já tem [IDEAL] ou [QUANT] curado
            if "[QUANT]" not in current_notas and key not in QUANT_DATA:
                new_nota = NEW_NOTAS[sid]
                # Preserve [SEM ACESSO] prefix if present
                if "[SEM ACESSO]" in current_notas:
                    continue
                # Replace generic notes
                if "[ABSTRACT]" in current_notas or current_notas == "" or \
                   not any(tag in current_notas for tag in ["[IDEAL]", "[QUALIT]", "[TEM_"]):
                    ws.cell(r, COL["Notas"]).value = sanitize(new_nota)
                    ws.cell(r, COL["Notas"]).fill = HIGHLIGHT_GREEN
                    updates["notas_new"] += 1

    # Salvar
    wb.save(tmp)
    wb.close()
    shutil.copy2(tmp, BD_PATH)
    print(f"\nSalvo: {BD_PATH}")

    print(f"\nAtualizações:")
    for k, v in updates.items():
        print(f"  {k:<15}: {v}")

    # Verificação final
    print(f"\n{'='*80}")
    print("COBERTURA FINAL")
    print("=" * 80)

    wb2 = openpyxl.load_workbook(tmp, read_only=True)
    ws2 = wb2.active
    total = ws2.max_row - 1

    for col_name in ["Proxy", "n_T", "m_T", "sd_T", "n_C", "m_C", "sd_C",
                     "Tipo_Intervencao", "Regiao", "Tipo_Comunidade", "Notas"]:
        ci = COL[col_name]
        filled = sum(1 for r in range(2, ws2.max_row + 1)
                     if ws2.cell(r, ci).value is not None and str(ws2.cell(r, ci).value).strip())
        print(f"  {col_name:<25}: {filled:3d}/{total} ({filled/total*100:.1f}%)")

    # Counts excluding sem acesso
    n_sem = len(IDS_SEM_ACESSO) * 6
    n_com = total - n_sem
    n_quant = sum(1 for r in range(2, ws2.max_row + 1)
                  if ws2.cell(r, COL["n_T"]).value is not None and str(ws2.cell(r, COL["n_T"]).value).strip())
    print(f"\n  Linhas totais: {total}")
    print(f"  Linhas com acesso a PDF: {n_com} ({n_com/total*100:.0f}%)")
    print(f"  Linhas sem acesso: {n_sem}")
    print(f"  n_T preenchido: {n_quant}/{total} total, {n_quant}/{n_com} com acesso ({n_quant/n_com*100:.1f}%)")

    # List all IDs with QUANT data
    ids_quant = set()
    for r in range(2, ws2.max_row + 1):
        if ws2.cell(r, COL["n_T"]).value is not None and str(ws2.cell(r, COL["n_T"]).value).strip():
            ids_quant.add(ws2.cell(r, COL["Study_ID"]).value)
    print(f"  Estudos com dados mean±SD: IDs {sorted(ids_quant)} ({len(ids_quant)} estudos)")

    wb2.close()


if __name__ == "__main__":
    main()
