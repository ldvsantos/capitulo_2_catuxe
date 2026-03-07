#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
40_recodificar_V1_V8.py
========================
Master recodification script: converts the 6-dimension bank (V1–V6) to the
8-dimension architecture (V1–V8) with condensation V2+V3→V2 (single dimension,
two proxy observations per study) and addition of V6 (Vitalidade Linguística),
V7 (Integração ao Mercado), V8 (Exposição Climática).

Mapping:
  OLD V1 → NEW V1  (Erosão Intergeracional e Migração)
  OLD V2 → NEW V2  (Complexidade e Singularidade Biocultural, proxy 1)
  OLD V3 → NEW V2  (Complexidade e Singularidade Biocultural, proxy 2)
  OLD V4 → NEW V3  (Status de Documentação)
  OLD V5 → NEW V4  (Vulnerabilidade Jurídica e Fundiária)
  OLD V6 → NEW V5  (Organização Social e Governança)
  ------- NEW V6   (Vitalidade Linguística)
  ------- NEW V7   (Integração ao Mercado)
  ------- NEW V8   (Exposição Climática)

Pipeline:
  1. Reads bd_extracao_PREENCHIDO.xlsx (288 rows, 48 studies × 6 dims)
  2. Remaps dimension codes
  3. Adds template rows for V6/V7/V8 per study (48 × 3 = 144 new rows)
  4. Auto-scans Notas field for keywords → pre-codes direction/intensity
  5. Saves bd_extracao_PREENCHIDO_V8.xlsx (432 rows)
  6. Updates bd_codificacao_qualitativa.xlsx with same remapping + auto-coding
  7. Produces diagnostic report

Author: Diego Vidal | 2026-03-06
"""

import os
import re
import copy
from collections import defaultdict, Counter

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
DIR_DADOS = os.path.join(BASE, "..", "2-BANCO_DADOS", "2-DADOS_TABULADOS")

FP_PREENCHIDO = os.path.join(DIR_DADOS, "bd_extracao_PREENCHIDO.xlsx")
FP_CODIFICACAO = os.path.join(DIR_DADOS, "bd_codificacao_qualitativa.xlsx")

FP_OUT_PREENCHIDO = os.path.join(DIR_DADOS, "bd_extracao_PREENCHIDO_V8.xlsx")
FP_OUT_CODIFICACAO = os.path.join(DIR_DADOS, "bd_codificacao_qualitativa_V8.xlsx")
FP_DIAGNOSTICO = os.path.join(DIR_DADOS, "diagnostico_recodificacao_V8.txt")

# ── Dimension mapping ────────────────────────────────────────────
DIM_MAP = {
    "V1": "V1",
    "V2": "V2",
    "V3": "V2",
    "V4": "V3",
    "V5": "V4",
    "V6": "V5",
}

LABEL_MAP = {
    "V1":  "Erosão Intergeracional e Migração",
    "V2":  "Complexidade e Singularidade Biocultural",
    "V3":  "Status de Documentação",
    "V4":  "Vulnerabilidade Jurídica e Fundiária",
    "V5":  "Organização Social e Governança",
    "V6":  "Vitalidade Linguística",
    "V7":  "Integração ao Mercado",
    "V8":  "Exposição Climática",
}

NEW_DIMS = ["V6", "V7", "V8"]

DEFAULT_PROXY = {
    "V6": "Linguistic vitality / ethnolinguistic indicator",
    "V7": "Market integration / commercialization indicator",
    "V8": "Climate exposure / extreme events indicator",
}

# ── Keyword dictionaries for auto-coding ─────────────────────────
# Each dict has 'vuln_increase' (direction +1) and 'vuln_decrease' (-1)
# keywords. Match is case-insensitive. Multi-word patterns allow fuzzy match.

KW_V6 = {
    "vuln_increase": [
        r"language\s+loss", r"language\s+shift", r"language\s+erosion",
        r"linguistic\s+erosion", r"endangered\s+language", r"language\s+death",
        r"monolingual\b", r"lingua\s+franca\s+replac",
        r"loss\s+of\s+(?:indigenous|local|native)\s+language",
        r"declining\s+(?:number\s+of\s+)?speakers",
        r"(?:perda|erosão)\s+linguíst", r"língua\s+(?:ameaçada|em\s+risco)",
        r"terminolog(?:y|ical)\s+(?:loss|erosion|decline)",
        r"etnotaxon[oô]mi\w*\s+(?:perda|erosão|declínio)",
    ],
    "vuln_decrease": [
        r"language\s+revitalization", r"linguistic\s+diversity\s+preserv",
        r"language\s+maintenance", r"language\s+(?:preservation|conservation)",
        r"bilingual\s+education", r"indigenous\s+language\s+program",
        r"(?:revitalização|preservação)\s+linguíst",
    ],
    "neutral": [
        r"\blanguage\b", r"\blinguistic\b", r"\bvernacular\b",
        r"\bethnolinguis", r"\betnotaxon", r"\bfolk\s+taxon",
        r"\bplant\s+names?\b", r"\blocal\s+names?\b",
        r"\bmother\s+tongue\b", r"\bindigenous\s+language\b",
        r"\blíngua\b", r"\bidioma\b", r"\bnomenc",
        r"\bEGIDS\b", r"\bVITEK\b",
    ],
}

KW_V7 = {
    "vuln_increase": [
        r"market\s+(?:integration|penetration|orientation)\s+(?:increas|expand|grow)",
        r"(?:cash|commercial)\s+crop\s+(?:replac|substitut|dominat)",
        r"(?:replacement|substitution)\s+of\s+(?:local|traditional|indigenous)\s+variet",
        r"(?:improved|hybrid|commercial)\s+variet(?:y|ies)\s+(?:adopt|replac|displac)",
        r"(?:abandono|substituição)\s+de\s+variedades?\s+(?:crioulas?|locais?|tradicionais?)",
        r"non[\s-]?farm\s+income\s+(?:increas|dominat)",
        r"market\s+(?:pressure|dependenc)",
        r"(?:loss|decline)\s+of\s+(?:subsistence|traditional)\s+(?:farming|agriculture)",
        r"(?:integração|pressão)\s+(?:de|do)\s+mercado",
    ],
    "vuln_decrease": [
        r"market\s+access\s+(?:benefit|improv|support)",
        r"fair\s+trade", r"(?:geographic|geographical)\s+indication",
        r"(?:valor|added\s+value)\s+(?:local|traditional)\s+products?",
        r"(?:niche|specialty)\s+market",
        r"(?:indicação\s+geográfica|marca\s+coletiva)",
    ],
    "neutral": [
        r"\bmarket\b", r"\bcommercial\b", r"\bsale\b", r"\btrade\b",
        r"\bincome\b", r"\bsubsistence\b", r"\blivelihood\b",
        r"\bcash\s+crop", r"\bcommodit", r"\bmarket\s+(?:access|distance)",
        r"\bmercado\b", r"\bcomercializ", r"\brenda\b",
    ],
}

KW_V8 = {
    "vuln_increase": [
        r"climate\s+change\s+(?:impact|effect|threat|vulnerab)",
        r"(?:drought|flood|flooding)\s+(?:impact|effect|increas|frequen|sever)",
        r"(?:extreme|unpredictable)\s+(?:weather|climate|rainfall|event)",
        r"(?:crop|harvest)\s+(?:failure|loss)\s+(?:due\s+to|caused\s+by)\s+(?:climate|drought|flood)",
        r"(?:seca|enchente|inundação)\s+(?:impacto|efeito|frequên)",
        r"(?:mudança|variabilidade)\s+climática\s+(?:impacto|ameaça|vulnerab)",
        r"(?:abandono|perda)\s+(?:de\s+)?(?:práticas?|cultivos?)\s+(?:por|devido)\s+(?:clima|seca)",
        r"(?:erratic|irregular)\s+(?:rainfall|precipitation)",
    ],
    "vuln_decrease": [
        r"climate\s+adaptation\s+(?:success|strateg|traditional)",
        r"traditional\s+(?:knowledge|practices?)\s+(?:for|in)\s+climate\s+(?:adaptation|resilience)",
        r"(?:indigenous|traditional|local)\s+(?:weather|climate)\s+(?:forecast|predict|indicator)",
        r"(?:adaptação|resiliência)\s+climática\s+(?:tradicional|indígena|local)",
    ],
    "neutral": [
        r"\bclimate\b", r"\bdrought\b", r"\bflood", r"\brainfall\b",
        r"\btemperature\b", r"\bprecipitation\b", r"\bweather\b",
        r"\bSPI\b", r"\bENSO\b", r"\bEl\s+Ni[ñn]o\b",
        r"\bseca\b", r"\benchente\b", r"\bpluvio",
        r"\bclimátic", r"\bseasonal",
        r"\bagroecological\s+zone",
    ],
}

KEYWORD_DICTS = {"V6": KW_V6, "V7": KW_V7, "V8": KW_V8}


def scan_text_for_dimension(text, kw_dict):
    """Scan text for keyword matches. Returns (direction, intensity, evidence_count, matched_keywords)."""
    if not text:
        return (None, None, 0, [])

    text_lower = text.lower()
    matches_pos = []
    matches_neg = []
    matches_neut = []

    for pattern in kw_dict["vuln_increase"]:
        found = re.findall(pattern, text_lower)
        if found:
            matches_pos.extend(found)

    for pattern in kw_dict["vuln_decrease"]:
        found = re.findall(pattern, text_lower)
        if found:
            matches_neg.extend(found)

    for pattern in kw_dict["neutral"]:
        found = re.findall(pattern, text_lower)
        if found:
            matches_neut.extend(found)

    total_directional = len(matches_pos) + len(matches_neg)
    total_all = total_directional + len(matches_neut)

    if total_all == 0:
        return (None, None, 0, [])

    # Determine direction
    if len(matches_pos) > len(matches_neg):
        direction = 1  # vulnerability increased
    elif len(matches_neg) > len(matches_pos):
        direction = -1  # vulnerability decreased
    elif total_directional > 0:
        direction = 0  # mixed evidence
    else:
        direction = 0  # only neutral terms, no directional evidence

    # Intensity: conservative
    if total_directional >= 3:
        intensity = 2  # moderate
    elif total_directional >= 1:
        intensity = 1  # weak
    else:
        intensity = None  # no directional evidence, only neutral

    all_matches = matches_pos + matches_neg + matches_neut
    truncated = [m[:50] for m in all_matches[:10]]

    return (direction, intensity, total_all, truncated)


# ══════════════════════════════════════════════════════════════════
# 1. RECODIFY bd_extracao_PREENCHIDO.xlsx
# ══════════════════════════════════════════════════════════════════

def recodify_preenchido():
    """Recodify the main extraction database."""
    print("=" * 70)
    print("STEP 1: Recodifying bd_extracao_PREENCHIDO.xlsx")
    print("=" * 70)

    if not os.path.exists(FP_PREENCHIDO):
        raise FileNotFoundError(f"File not found: {FP_PREENCHIDO}")

    wb = openpyxl.load_workbook(FP_PREENCHIDO)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    print(f"  Columns: {headers}")

    # Column indices
    idx = {h: i for i, h in enumerate(headers)}
    dim_col = idx["Dimensao"]
    label_col = idx["Dimensao_Label"]
    proxy_col = idx.get("Proxy", idx.get("Proxy", None))
    notas_col = idx.get("Notas", None)
    tier_col = idx.get("Tier", None)
    dir_col = idx.get("Direcao_efeito", None)
    int_col = idx.get("Intensidade", None)
    conf_col = idx.get("Confianca_codificacao", None)
    notes_cod_col = idx.get("Notas_codificador", None)
    rev_col = idx.get("Revisor", None)

    # Read all data rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))

    print(f"  Original rows: {len(rows)}")

    # Group by Study_ID to get study metadata templates
    study_meta = {}
    for row in rows:
        sid = row[idx["Study_ID"]]
        if sid not in study_meta:
            study_meta[sid] = row.copy()

    print(f"  Unique studies: {len(study_meta)}")

    # Step A: Remap existing dimensions
    remap_count = Counter()
    for row in rows:
        old_dim = row[dim_col]
        if old_dim in DIM_MAP:
            new_dim = DIM_MAP[old_dim]
            row[dim_col] = new_dim
            row[label_col] = LABEL_MAP.get(new_dim, row[label_col])
            remap_count[f"{old_dim}→{new_dim}"] += 1

    print(f"  Remapped: {dict(remap_count)}")

    # Step B: Add V6, V7, V8 rows for each study
    new_rows = []
    auto_code_stats = Counter()

    for sid in sorted(study_meta.keys()):
        template = study_meta[sid]
        notas_text = template[notas_col] if notas_col is not None else ""

        for new_dim in NEW_DIMS:
            new_row = [None] * len(headers)
            # Copy metadata columns
            for col_name in ["Study_ID", "Study", "DOI", "Title", "Journal", "Year",
                             "Tipo_Intervencao", "Regiao", "Tipo_Comunidade",
                             "Tempo_Intervencao"]:
                if col_name in idx:
                    new_row[idx[col_name]] = template[idx[col_name]]

            # Set dimension info
            new_row[dim_col] = new_dim
            new_row[label_col] = LABEL_MAP[new_dim]
            if proxy_col is not None:
                new_row[proxy_col] = DEFAULT_PROXY[new_dim]

            # Copy Notas (abstract/extracted text) for reference
            if notas_col is not None:
                new_row[notas_col] = notas_text

            # Quantitative fields remain None (n_T, m_T, sd_T, n_C, m_C, sd_C, NOS)

            # Auto-code from keywords
            kw_dict = KEYWORD_DICTS[new_dim]
            direction, intensity, n_matches, matched = scan_text_for_dimension(
                notas_text, kw_dict
            )

            if tier_col is not None:
                new_row[tier_col] = "T4" if direction is not None else None
            if dir_col is not None:
                new_row[dir_col] = direction
            if int_col is not None:
                new_row[int_col] = intensity
            if conf_col is not None:
                if n_matches > 0:
                    new_row[conf_col] = "AUTO"
                else:
                    new_row[conf_col] = None
            if notes_cod_col is not None:
                if matched:
                    new_row[notes_cod_col] = f"[AUTO] {n_matches} matches: {'; '.join(matched[:5])}"
                else:
                    new_row[notes_cod_col] = "[AUTO] No keyword matches found"

            new_rows.append(new_row)

            # Stats
            if direction is not None:
                auto_code_stats[f"{new_dim}_coded"] += 1
            else:
                auto_code_stats[f"{new_dim}_no_evidence"] += 1

    print(f"  New rows added: {len(new_rows)}")
    print(f"  Auto-coding stats: {dict(auto_code_stats)}")

    # Combine original + new rows
    all_rows = rows + new_rows

    # Sort by Study_ID, then dimension
    dim_order = ["V1", "V2", "V3", "V4", "V5", "V6", "V7", "V8"]
    dim_sort_key = {d: i for i, d in enumerate(dim_order)}

    all_rows.sort(key=lambda r: (
        r[idx["Study_ID"]] if r[idx["Study_ID"]] is not None else 999,
        dim_sort_key.get(r[dim_col], 99)
    ))

    # Write output
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sheet1"

    # Header
    for i, h in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

    # Styles for new rows
    fill_new = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    fill_auto = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    for r_idx, row in enumerate(all_rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws_out.cell(row=r_idx, column=c_idx, value=val)
            # Highlight new dimensions
            dim_val = row[dim_col]
            if dim_val in NEW_DIMS:
                if row[conf_col] == "AUTO" if conf_col is not None else False:
                    cell.fill = fill_auto
                else:
                    cell.fill = fill_new

    # Auto-fit column widths (approximate)
    for col in range(1, len(headers) + 1):
        ws_out.column_dimensions[get_column_letter(col)].width = 18

    wb_out.save(FP_OUT_PREENCHIDO)
    print(f"  ✔ Saved: {FP_OUT_PREENCHIDO}")
    print(f"  Total rows: {len(all_rows)}")

    return all_rows, headers, idx, auto_code_stats


# ══════════════════════════════════════════════════════════════════
# 2. RECODIFY bd_codificacao_qualitativa.xlsx
# ══════════════════════════════════════════════════════════════════

def recodify_codificacao():
    """Recodify the qualitative coding database."""
    print("\n" + "=" * 70)
    print("STEP 2: Recodifying bd_codificacao_qualitativa.xlsx")
    print("=" * 70)

    if not os.path.exists(FP_CODIFICACAO):
        print("  ⚠ File not found, skipping codificacao recodification.")
        return None

    wb = openpyxl.load_workbook(FP_CODIFICACAO)
    ws = wb.active
    headers_cod = [c.value for c in ws[1]]
    print(f"  Columns: {headers_cod}")

    idx_c = {h: i for i, h in enumerate(headers_cod)}
    dim_col_c = idx_c.get("Dimensao", None)
    label_col_c = idx_c.get("Dimensao_Label", None)

    if dim_col_c is None:
        print("  ⚠ 'Dimensao' column not found, skipping.")
        return None

    # Read rows
    rows_cod = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows_cod.append(list(row))

    print(f"  Original rows: {len(rows_cod)}")

    # The codificacao has Dimensao in a different position (it's in index 4 based on earlier output)
    # Actually let me check: the Dimensao column is at index dim_col_c
    # But the issue is that in codificacao, the Dimensao column contains Proxy names, not dimension codes!
    # Let me check the actual column content...

    # From the earlier terminal output, the codificacao "Dimensao" column actually stores
    # proxy names like "Agrobiodiversidade", "Complexidade Cultural", etc.
    # And "Dimensao_Label" stores... let me check
    # Actually from the output: columns are:
    # ['Study_ID', 'Study', 'DOI', 'Year', 'Dimensao', 'Dimensao_Label', 'Proxy', ...]
    # And the counts showed things like "Agrobiodiversidade: 15 registros" — that was from
    # the Proxy column (index 6), not Dimensao (index 4).

    # Let me just remap Dimensao and Dimensao_Label
    remap_count_c = Counter()

    # Check what values are in the Dimensao column
    dim_values = set()
    for row in rows_cod:
        if dim_col_c < len(row) and row[dim_col_c]:
            dim_values.add(row[dim_col_c])

    print(f"  Dimensao values found: {sorted(dim_values)}")

    # Remap if values are V1-V6 style
    v_style = any(v in dim_values for v in ["V1", "V2", "V3", "V4", "V5", "V6"])

    if v_style:
        for row in rows_cod:
            old_dim = row[dim_col_c]
            if old_dim in DIM_MAP:
                new_dim = DIM_MAP[old_dim]
                row[dim_col_c] = new_dim
                if label_col_c is not None and label_col_c < len(row):
                    row[label_col_c] = LABEL_MAP.get(new_dim, row[label_col_c])
                remap_count_c[f"{old_dim}→{new_dim}"] += 1
        print(f"  Remapped: {dict(remap_count_c)}")
    else:
        print("  Dimensao values are not V1-V6 style. Checking Dimensao_Label...")
        # Try mapping based on Dimensao_Label
        label_to_old_dim = {
            "Erosao Intergeracional": "V1", "Erosão Intergeracional": "V1",
            "Complexidade Biocultural": "V2", "Complexidade Cultural": "V2",
            "Singularidade Territorial": "V3",
            "Status de Documentacao": "V4", "Status de Documentação": "V4",
            "Vulnerabilidade Juridica": "V5", "Vulnerabilidade Jurídica": "V5",
            "Organizacao Social": "V6", "Organização Social": "V6",
        }
        for row in rows_cod:
            label_val = row[label_col_c] if label_col_c is not None and label_col_c < len(row) else None
            if label_val:
                for lbl, old_d in label_to_old_dim.items():
                    if lbl.lower() in str(label_val).lower():
                        new_d = DIM_MAP[old_d]
                        row[dim_col_c] = new_d
                        row[label_col_c] = LABEL_MAP.get(new_d, label_val)
                        remap_count_c[f"{old_d}→{new_d}"] += 1
                        break
        print(f"  Remapped via labels: {dict(remap_count_c)}")

    # Write output
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "CODIFICACAO"

    for i, h in enumerate(headers_cod, 1):
        cell = ws_out.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True)

    for r_idx, row in enumerate(rows_cod, 2):
        for c_idx, val in enumerate(row, 1):
            ws_out.cell(row=r_idx, column=c_idx, value=val)

    wb_out.save(FP_OUT_CODIFICACAO)
    print(f"  ✔ Saved: {FP_OUT_CODIFICACAO}")

    return rows_cod


# ══════════════════════════════════════════════════════════════════
# 3. DIAGNOSTIC REPORT
# ══════════════════════════════════════════════════════════════════

def write_diagnostico(all_rows, headers, idx, auto_stats):
    """Write a diagnostic report."""
    print("\n" + "=" * 70)
    print("STEP 3: Generating diagnostic report")
    print("=" * 70)

    dim_col = idx["Dimensao"]
    sid_col = idx["Study_ID"]
    dir_col = idx.get("Direcao_efeito", None)
    conf_col = idx.get("Confianca_codificacao", None)

    lines = []
    lines.append("DIAGNOSTIC REPORT: Recodification V1-V6 → V1-V8")
    lines.append("=" * 60)
    lines.append(f"Total rows: {len(all_rows)}")

    # Count by dimension
    dim_ct = Counter()
    study_ct_per_dim = defaultdict(set)
    coded_ct = Counter()

    for row in all_rows:
        d = row[dim_col]
        dim_ct[d] += 1
        study_ct_per_dim[d].add(row[sid_col])
        if dir_col is not None and row[dir_col] is not None:
            coded_ct[d] += 1

    lines.append("\nDimension | Rows | Studies | Auto-coded")
    lines.append("-" * 50)
    dim_order = ["V1", "V2", "V3", "V4", "V5", "V6", "V7", "V8"]
    for d in dim_order:
        n = dim_ct.get(d, 0)
        k = len(study_ct_per_dim.get(d, set()))
        c = coded_ct.get(d, 0)
        label = LABEL_MAP.get(d, "?")
        lines.append(f"  {d:4s} | {n:4d} | {k:4d}    | {c:4d}  ({label})")

    # Auto-coding summary for V6-V8
    lines.append("\nAuto-coding summary (V6-V8):")
    lines.append("-" * 50)
    for d in NEW_DIMS:
        coded = auto_stats.get(f"{d}_coded", 0)
        no_ev = auto_stats.get(f"{d}_no_evidence", 0)
        total = coded + no_ev
        pct = (coded / total * 100) if total > 0 else 0
        lines.append(f"  {d} ({LABEL_MAP[d]}): {coded}/{total} auto-coded ({pct:.0f}%)")

    # Direction distribution for auto-coded V6-V8
    lines.append("\nDirection distribution (auto-coded V6-V8):")
    lines.append("-" * 50)
    for d in NEW_DIMS:
        dir_dist = Counter()
        for row in all_rows:
            if row[dim_col] == d and dir_col is not None and row[dir_col] is not None:
                dir_dist[row[dir_col]] += 1
        lines.append(f"  {d}: +1={dir_dist.get(1,0)}, 0={dir_dist.get(0,0)}, -1={dir_dist.get(-1,0)}")

    lines.append("\n" + "=" * 60)
    lines.append("NEXT STEPS:")
    lines.append("  1. Review auto-coded V6-V8 rows (yellow/green highlighting)")
    lines.append("  2. Manually validate direction/intensity for flagged rows")
    lines.append("  3. Run 30_preparar_codificacao_quali.py on _V8 file")
    lines.append("  4. Run 31_converter_quali_para_lnRR.py")
    lines.append("  5. Run R pipeline (02b → 03 → 04 → 05 → 06)")
    lines.append("=" * 60)

    report = "\n".join(lines)
    with open(FP_DIAGNOSTICO, "w", encoding="utf-8") as f:
        f.write(report)

    print(report)
    print(f"\n  ✔ Report saved: {FP_DIAGNOSTICO}")


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    all_rows, headers, idx, auto_stats = recodify_preenchido()
    recodify_codificacao()
    write_diagnostico(all_rows, headers, idx, auto_stats)
    print("\n✔ Recodification complete.")
