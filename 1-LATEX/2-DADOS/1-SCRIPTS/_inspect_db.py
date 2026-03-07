import openpyxl
wb = openpyxl.load_workbook('2-BANCO_DADOS/2-DADOS_TABULADOS/bd_extracao_PREENCHIDO_V8.xlsx')
ws = wb.active
headers = [c.value for c in ws[1]]
print('COLUNAS:', headers)
print()

rows_by_dim = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    r = dict(zip(headers, row))
    dim = r.get('Dimensao')
    if dim not in rows_by_dim:
        rows_by_dim[dim] = []
    rows_by_dim[dim].append(r)

# V1 coded examples
print('=== EXEMPLOS V1 (codificados) ===')
for r in rows_by_dim.get('V1', [])[:5]:
    if r.get('Direcao_efeito') is not None:
        nc = str(r.get('Notas_codificador', '') or '')[:80]
        print(f"Study={r['Study_ID']}, Tier={r['Tier']}, Dir={r['Direcao_efeito']}, Int={r['Intensidade']}, Conf={r['Confianca_codificacao']}, NotasCod={nc}")

print()
print('=== EXEMPLOS V4 (codificados) ===')
for r in rows_by_dim.get('V4', [])[:5]:
    if r.get('Direcao_efeito') is not None:
        nc = str(r.get('Notas_codificador', '') or '')[:80]
        print(f"Study={r['Study_ID']}, Tier={r['Tier']}, Dir={r['Direcao_efeito']}, Int={r['Intensidade']}, Conf={r['Confianca_codificacao']}, NotasCod={nc}")

print()
for dim in ['V6', 'V7', 'V8']:
    print(f'=== EXEMPLOS {dim} (template) ===')
    for r in rows_by_dim.get(dim, [])[:4]:
        nc = str(r.get('Notas_codificador', '') or '')[:60]
        nt = str(r.get('Notas', '') or '')[:60]
        print(f"Study={r['Study_ID']}, Tier={r['Tier']}, Dir={r['Direcao_efeito']}, Int={r['Intensidade']}, Conf={r['Confianca_codificacao']}, NotasCod={nc}, Notas={nt}")
    print()

# Also list all 48 studies for V6 with their Study + Title
print('=== TODOS OS 48 ESTUDOS (V6 como referência) ===')
for i, r in enumerate(rows_by_dim.get('V6', [])):
    title = str(r.get('Title', '') or '')[:70]
    study = str(r.get('Study', '') or '')[:30]
    print(f"{i+1}. ID={r['Study_ID']}, {study} | {title}")
