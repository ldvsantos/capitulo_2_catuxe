# -*- coding: utf-8 -*-
"""
19_listar_sem_pdf.py
====================
Lista os 22 estudos sem PDF e extrai DOIs para busca no Scopus.
"""
import openpyxl, shutil, tempfile, sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

BASE = Path(r"c:\Users\vidal\OneDrive\Documentos\13 - CLONEGIT"
            r"\artigo_2_catuxe\1-LATEX\2-DADOS\2-BANCO_DADOS")
TAB = BASE / "2-DADOS_TABULADOS"

# Copiar para temp (evitar lock do Excel)
tmp_bd = Path(tempfile.gettempdir()) / "bd_extracao_copy.xlsx"
tmp_sel = Path(tempfile.gettempdir()) / "sel42_copy.xlsx"
shutil.copy2(TAB / "bd_extracao.xlsx", tmp_bd)
shutil.copy2(TAB / "selecionados_42_completos.xlsx", tmp_sel)

# ── IDs com PDF (mapeamento existente) ──
PDF_TO_BDID = {
    "18-doyle": 36, "A-Biodiversity-Hotspot": 3, "Beyond-biodiversity": 39,
    "Biocultural-conservation-systems": 13, "Colourful-agrobiodiversity": 4,
    "Comparison-of-medicinal": 9, "Erosion-of-traditional": 15,
    "Ethnobotany-and-conservation": 16, "Ethnobotany-of-the-Aegadian": 27,
    "Exploring-farmers": 11, "Guardians-of-heritage": 20,
    "Interconnected-Nature": 19, "Mountain-Graticules": 6,
    "Stingless-bee-keeping": 7, "Strategies-for-managing": 5,
    "Towards-biocultural": 14, "Unearthing-Unevenness": 37,
    "Voices-Around-the-South": 30, "ajol-file-journals": 32,
    "The-Zo-perspective": 8,
    "1-s2.0-S1470160X20308037": 43, "s12231-018-9401-y": 44,
    "A-Quantitative-Study": 45, "1-s2.0-S1462901124001953": 46,
    "Socialecological": 47, "s10722-017-0544-y": 48,
}
ids_com_pdf = set(PDF_TO_BDID.values())
all_ids = set(range(1, 49))
ids_sem_pdf = sorted(all_ids - ids_com_pdf)

# ── Ler bd_extracao para autores/titulos ──
wb = openpyxl.load_workbook(tmp_bd, read_only=True)
ws = wb.active
hdr_bd = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
print("BD COLUNAS:", hdr_bd[:8])

seen = set()
estudos_sem_pdf = []
for row in ws.iter_rows(min_row=2, values_only=True):
    vals = list(row)
    sid = vals[0]
    if sid in ids_sem_pdf and sid not in seen:
        seen.add(sid)
        estudos_sem_pdf.append((sid, vals[1], vals[2], vals[3]))
wb.close()

# ── Ler selecionados_42 para DOIs ──
wb2 = openpyxl.load_workbook(tmp_sel, read_only=True)
ws2 = wb2.active
hdr2 = [c.value for c in next(ws2.iter_rows(min_row=1, max_row=1))]
print("SEL42 COLUNAS:", [str(h)[:30] for h in hdr2])

doi_col = None
title_col = None
author_col = None
for i, h in enumerate(hdr2):
    hl = str(h).lower() if h else ""
    if "doi" in hl:
        doi_col = i
    if "title" in hl or "titulo" in hl:
        if title_col is None:
            title_col = i
    if "author" in hl or "autor" in hl:
        if author_col is None:
            author_col = i

print(f"DOI col={doi_col}, Title col={title_col}, Author col={author_col}")

# Map title->DOI
title_doi = {}
author_doi = {}
for row in ws2.iter_rows(min_row=2, values_only=True):
    vals = list(row)
    t = str(vals[title_col]).strip().lower() if title_col is not None and vals[title_col] else ""
    d = str(vals[doi_col]).strip() if doi_col is not None and vals[doi_col] else ""
    a = str(vals[author_col]).strip().lower() if author_col is not None and vals[author_col] else ""
    if t and d:
        title_doi[t] = d
    if a and d:
        # Use first author surname
        surname = a.split(",")[0].split(" ")[0].strip()
        if surname:
            author_doi[surname] = d
wb2.close()

print(f"\nTotal selecionados_42: {len(title_doi)} entries com DOI")
print(f"IDs sem PDF: {ids_sem_pdf}")
print(f"Total sem PDF: {len(ids_sem_pdf)}")
print()

# ── Match e listar ──
print("=" * 130)
fmt = "{:>3}  {:<25} {:>4}  {:<50}  {}"
print(fmt.format("ID", "Autor", "Ano", "DOI", "Titulo"))
print("-" * 130)

dois_found = []
no_doi = []

for sid, autor, ano, titulo in sorted(estudos_sem_pdf):
    t_lower = str(titulo).strip().lower() if titulo else ""
    a_lower = str(autor).strip().lower() if autor else ""
    
    # Try exact title match
    doi = title_doi.get(t_lower, "")
    
    # Try fuzzy title match (first 40 chars)
    if not doi and t_lower:
        for tk, dv in title_doi.items():
            if t_lower[:40] in tk or tk[:40] in t_lower:
                doi = dv
                break
    
    # Try author surname match
    if not doi and a_lower:
        surname = a_lower.split(",")[0].split(" ")[0].strip()
        doi = author_doi.get(surname, "")
    
    if doi:
        dois_found.append(doi)
    else:
        no_doi.append((sid, autor, titulo))
    
    print(fmt.format(sid, str(autor)[:25], str(ano)[:4], doi[:50], str(titulo)[:55]))

print(f"\nCom DOI: {len(dois_found)}")
print(f"Sem DOI: {len(no_doi)}")
if no_doi:
    print("\nEstudos sem DOI encontrado:")
    for sid, autor, titulo in no_doi:
        print(f"  ID={sid}: {autor} - {titulo}")

# ── Gerar string Scopus ──
print("\n\n" + "=" * 130)
print("STRING PARA BUSCA NO SCOPUS (colar direto no campo de busca)")
print("=" * 130)

# Scopus aceita: DOI("10.xxxx/yyyy") OR DOI("10.xxxx/zzzz")
if dois_found:
    parts = []
    for d in dois_found:
        # Limpar DOI
        d = d.strip()
        if d.startswith("http"):
            d = d.split("doi.org/")[-1]
        if d:
            parts.append(f'DOI("{d}")')
    
    scopus_str = " OR ".join(parts)
    print(f"\n{scopus_str}")
    print(f"\n\nTotal DOIs na string: {len(parts)}")
else:
    print("\nNenhum DOI encontrado!")

# Se houver estudos sem DOI, gerar busca por título
if no_doi:
    print("\n\n" + "=" * 130)
    print("BUSCA ALTERNATIVA POR TÍTULO (para estudos sem DOI)")
    print("=" * 130)
    for sid, autor, titulo in no_doi:
        if titulo:
            # Primeiras 8 palavras do título
            words = str(titulo).split()[:8]
            title_q = " ".join(words)
            print(f'\nID={sid}: TITLE("{title_q}")')
