#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Inspect all 48 studies for V6-V8 coding context."""
import openpyxl, os

fp = os.path.join(os.path.dirname(__file__),
                  '2-BANCO_DADOS', '2-DADOS_TABULADOS',
                  'bd_extracao_PREENCHIDO_V8.xlsx')
wb = openpyxl.load_workbook(fp)
ws = wb.active
hdrs = [c.value for c in ws[1]]
idx = {h: i for i, h in enumerate(hdrs)}

studies = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    sid = row[idx['Study_ID']]
    if sid not in studies:
        studies[sid] = {
            'study': row[idx['Study']],
            'title': row[idx['Title']],
            'journal': row[idx['Journal']],
            'year': row[idx['Year']],
            'notas': row[idx['Notas']],
            'regiao': row[idx['Regiao']],
            'tipo_com': row[idx['Tipo_Comunidade']],
            'dims': {}
        }
    dim = row[idx['Dimensao']]
    studies[sid]['dims'][dim] = {
        'dir': row[idx['Direcao_efeito']],
        'int': row[idx['Intensidade']],
        'tier': row[idx['Tier']],
        'conf': row[idx['Confianca_codificacao']],
    }

for sid in sorted(studies.keys()):
    s = studies[sid]
    study_name = s['study'] or '?'
    title = s['title'] or '?'
    journal = s['journal'] or '?'
    year = s['year'] or '?'
    regiao = s['regiao'] or '?'
    tipo_com = s['tipo_com'] or '?'
    
    print(f"=== STUDY {sid}: {study_name} ({year}) ===")
    print(f"  Title: {title}")
    print(f"  Journal: {journal} | Region: {regiao} | Community: {tipo_com}")
    
    # V1 coding as reference
    v1 = s['dims'].get('V1', {})
    print(f"  V1: Dir={v1.get('dir')} Int={v1.get('int')} Tier={v1.get('tier')} Conf={v1.get('conf')}")
    
    # V6-V8 current
    for vx in ['V6', 'V7', 'V8']:
        vd = s['dims'].get(vx, {})
        print(f"  {vx}: Dir={vd.get('dir')} Int={vd.get('int')} Tier={vd.get('tier')}")
    
    # Notas
    notas = str(s['notas'])[:400] if s['notas'] else '(none)'
    print(f"  Notas: {notas}")
    print()
