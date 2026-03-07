#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
31_converter_quali_para_lnRR.py
================================
Lê bd_codificacao_qualitativa.xlsx (após codificação manual dos revisores) e
converte evidência qualitativa/semi-quantitativa em tamanhos de efeito (lnRR)
com propagação de incerteza, seguindo o framework multi-tier.

Conversões implementadas:
  T1  — lnRR direto: ln(m_T/m_C), vi = sd_T²/(n_T·m_T²) + sd_C²/(n_C·m_C²)
  T2a — p-value → z → d → lnRR via CV mediano dos T1 da mesma dimensão
  T2b — Idem T2a, usando p-threshold como p-value conservador
  T3  — Ordinal (direção × intensidade) → quantil logit-normal → lnOR → lnRR
  T4  — Idem T3, com inflação adicional de variância

Referências metodológicas:
  - Hasselblad & Hedges (1995) para conversão lnOR ↔ d
  - Zhang & Yu (1998) para conversão lnOR → lnRR
  - Borenstein et al. (2009, Cap. 7) para d a partir de p-values
  - Regra delta para propagação de variância nas conversões

Saída: bd_extracao_convertido.xlsx (pronto para o pipeline R)
       bd_extracao_convertido.csv  (backup)

Autor: Diego Vidal  |  2026-03-02
"""

import math
import os
import re
import warnings
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill, Font

# ── Caminhos ──────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
DIR_DADOS = os.path.join(BASE, "..", "2-BANCO_DADOS", "2-DADOS_TABULADOS")
FP_COD = os.path.join(DIR_DADOS, "bd_codificacao_qualitativa_V8.xlsx")
FP_ORIG = os.path.join(DIR_DADOS, "bd_extracao_PREENCHIDO_V8.xlsx")
FP_OUT_XLSX = os.path.join(DIR_DADOS, "bd_extracao_convertido.xlsx")
FP_OUT_CSV = os.path.join(DIR_DADOS, "bd_extracao_convertido.csv")

# ── Constantes estatísticas ───────────────────────────────────────
PI = math.pi
PI2_OVER_3 = PI**2 / 3  # ≈ 3.290, variância da logística padrão
SQRT_PI2_3 = math.sqrt(PI2_OVER_3)  # ≈ 1.814

# Quantis da logit-normal para mapeamento ordinal (intensidade 1, 2, 3)
# Logit(p) = ln(p/(1-p)) para p = 0.40, 0.60, 0.80
LOGIT_MAP = {
    1: math.log(0.40 / 0.60),   # ≈ -0.405  (efeito fraco)
    2: math.log(0.60 / 0.40),   # ≈ +0.405  (efeito moderado)
    3: math.log(0.80 / 0.20),   # ≈ +1.386  (efeito forte)
}

# Variância fixa para lnOR ordinal (prior vago)
VI_ORDINAL_BASE = PI2_OVER_3  # ≈ 3.290

# Fator de inflação de variância para T4 (qualitativo puro, mais incerteza)
INFLATE_T4 = 1.5

# Média nT+nC de fallback quando n não disponível (mediana de campo)
N_FALLBACK = 60  # conservador para estudos etnobotânicos


def norm_ppf(p):
    """Inversa da CDF normal padrão (aproximação Abramowitz & Stegun 26.2.23)."""
    if p <= 0 or p >= 1:
        return float('inf') if p >= 1 else float('-inf')
    if p < 0.5:
        return -norm_ppf(1 - p)
    t = math.sqrt(-2 * math.log(1 - p))
    c0, c1, c2 = 2.515517, 0.802853, 0.010328
    d1, d2, d3 = 1.432788, 0.189269, 0.001308
    return t - (c0 + c1 * t + c2 * t**2) / (1 + d1 * t + d2 * t**2 + d3 * t**3)


def p_to_z(p_value, two_sided=True):
    """Converte p-value em z-score."""
    if two_sided:
        return norm_ppf(1 - p_value / 2)
    return norm_ppf(1 - p_value)


def d_from_z_and_n(z, n_T, n_C):
    """Hedges' d a partir de z e tamanhos amostrais.
    d = z * sqrt(1/n_T + 1/n_C)  (Borenstein et al. 2009, Eq. 7.5)
    """
    return z * math.sqrt(1/n_T + 1/n_C)


def vi_d(d, n_T, n_C):
    """Variância amostral de d.
    v_d = 1/n_T + 1/n_C + d²/(2*(n_T + n_C))  (Borenstein 2009, Eq. 7.6)
    """
    return 1/n_T + 1/n_C + d**2 / (2 * (n_T + n_C))


def d_to_lnOR(d):
    """Converte d (SMD) em lnOR via Hasselblad & Hedges (1995).
    lnOR = d * π / √3
    """
    return d * PI / math.sqrt(3)


def vi_d_to_vi_lnOR(v_d):
    """Converte variância de d para variância de lnOR.
    v_lnOR = v_d * π²/3  (regra delta)
    """
    return v_d * PI2_OVER_3


def lnOR_to_lnRR(lnOR, p0=0.5):
    """Converte lnOR em lnRR via Zhang & Yu (1998).
    lnRR = ln[ p0 * (exp(lnOR) - 1) + 1 ] - ln(p0 * exp(lnOR) + 1 - p0)
    
    Simplificação: quando p0 = 0.5 (prior não-informativo):
    lnRR ≈ lnOR * (1 - p0) para lnOR pequenos
    
    Forma exata:
    RR = OR / (1 - p0 + p0 * OR)
    lnRR = lnOR - ln(1 - p0 + p0 * exp(lnOR))
    """
    OR = math.exp(lnOR)
    RR = OR / (1 - p0 + p0 * OR)
    if RR <= 0:
        return 0.0
    return math.log(RR)


def vi_lnOR_to_vi_lnRR(vi_lnOR, lnOR, p0=0.5):
    """Variância de lnRR via regra delta sobre a conversão Zhang & Yu.
    ∂lnRR/∂lnOR = (1-p0) * exp(lnOR) / (1 - p0 + p0*exp(lnOR))
    v_lnRR = v_lnOR * (∂lnRR/∂lnOR)²
    """
    OR = math.exp(lnOR)
    denom = 1 - p0 + p0 * OR
    if denom == 0:
        return vi_lnOR
    deriv = (1 - p0) * OR / denom
    return vi_lnOR * deriv**2


def lnRR_direto(m_T, sd_T, n_T, m_C, sd_C, n_C):
    """Cálculo direto de lnRR e vi para dados quantitativos (Tier 1)."""
    if m_T <= 0 or m_C <= 0:
        return None, None
    lnrr = math.log(m_T / m_C)
    vi = (sd_T**2) / (n_T * m_T**2) + (sd_C**2) / (n_C * m_C**2)
    return lnrr, max(vi, 1e-6)


def converter_t2a(p_value, direction, n_T, n_C, cv_ref=None):
    """Converte p-value + direção + n em lnRR.
    
    Pipeline: p → z → d → lnOR → lnRR
    Se cv_ref disponível (CV mediano dos T1), usa conversão direta: lnRR ≈ d * cv_ref
    """
    z = p_to_z(p_value)
    d = d_from_z_and_n(z, n_T, n_C)
    d = d * direction  # aplica sinal

    v_d = vi_d(d, n_T, n_C)

    if cv_ref is not None and cv_ref > 0:
        # Conversão direta via CV (mais precisa quando temos referência)
        # lnRR ≈ d * CV_ref (Lajeunesse 2011)
        lnrr = d * cv_ref
        vi_lnrr = v_d * cv_ref**2
    else:
        # Conversão via lnOR → lnRR (Zhang & Yu 1998)
        lnor = d_to_lnOR(d)
        v_lnor = vi_d_to_vi_lnOR(v_d)
        lnrr = lnOR_to_lnRR(lnor, p0=0.5)
        vi_lnrr = vi_lnOR_to_vi_lnRR(v_lnor, lnor, p0=0.5)

    return lnrr, max(vi_lnrr, 1e-6)


def converter_ordinal(direction, intensity, n_T=None, n_C=None):
    """Converte codificação ordinal em lnRR.
    
    Pipeline: (direção, intensidade) → lnOR logit → lnRR (Zhang & Yu)
    Variância = π²/3 como prior vago (Hasselblad & Hedges 1995)
    """
    if direction == 0:
        return 0.0, VI_ORDINAL_BASE

    lnor = LOGIT_MAP.get(abs(intensity), LOGIT_MAP[2])  # default moderado
    lnor = lnor * direction  # aplica sinal

    v_lnor = VI_ORDINAL_BASE

    # Se n disponível, ajustar variância (mais preciso)
    if n_T and n_C and n_T > 0 and n_C > 0:
        # Variância empírica de lnOR com n: 1/a + 1/b + 1/c + 1/d
        # Aproximação: v_lnOR ≈ 1/(n_T*p_T*(1-p_T)) + 1/(n_C*p_C*(1-p_C))
        # Com p desconhecido, usar π²/3 mas ponderar por n
        v_lnor = min(v_lnor, 4.0 / (n_T + n_C) + v_lnor * 0.5)

    lnrr = lnOR_to_lnRR(lnor, p0=0.5)
    vi_lnrr = vi_lnOR_to_vi_lnRR(v_lnor, lnor, p0=0.5)

    return lnrr, max(vi_lnrr, 1e-6)


def calcular_cv_por_dimensao(registros_t1):
    """Calcula CV mediano dos estudos T1 por dimensão, para calibrar conversões."""
    cvs = defaultdict(list)
    for r in registros_t1:
        if r["m_C"] and r["sd_C"] and r["m_C"] > 0:
            cv = r["sd_C"] / r["m_C"]
            cvs[r["Dimensao"]].append(cv)
    cv_mediano = {}
    for dim, vals in cvs.items():
        vals.sort()
        mid = len(vals) // 2
        cv_mediano[dim] = vals[mid] if vals else None
    return cv_mediano


def safe_float(val):
    """Converte para float, retorna None se inválido."""
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_int(val):
    """Converte para int, retorna None se inválido."""
    if val is None or val == "":
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def main():
    # ── 1. Ler registros T1 do arquivo original ──
    wb_orig = openpyxl.load_workbook(FP_ORIG, read_only=True)
    ws_orig = wb_orig["Sheet1"]

    # IDs excluídos na fase de codificação
    IDS_EXCLUIDOS = {36}

    registros_t1 = []
    for row in ws_orig.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        # Filtrar estudos excluídos
        if safe_int(row[0]) in IDS_EXCLUIDOS:
            continue
        n_T, m_T, sd_T = row[9], row[10], row[11]
        n_C, m_C, sd_C = row[12], row[13], row[14]
        if not all(v is not None for v in [n_T, m_T, sd_T, n_C, m_C, sd_C]):
            continue
        lnrr, vi = lnRR_direto(m_T, sd_T, n_T, m_C, sd_C, n_C)
        registros_t1.append({
            "Study_ID": row[0], "Study": row[1], "DOI": row[2],
            "Year": row[5], "Dimensao": row[6], "Dimensao_Label": row[7],
            "Proxy": row[8],
            "n_T": n_T, "m_T": m_T, "sd_T": sd_T,
            "n_C": n_C, "m_C": m_C, "sd_C": sd_C,
            "Tipo_Intervencao": row[15], "Regiao": row[16],
            "Tipo_Comunidade": row[17], "NOS": row[19],
            "Tier": "T1", "lnRR": lnrr, "vi": vi,
            "sigma_conv": 0.0,  # sem conversão
        })
    wb_orig.close()

    # CV mediano por dimensão (para calibrar conversão T2)
    cv_dim = calcular_cv_por_dimensao(registros_t1)
    print("CV medianos por dimensao (referencia T1):")
    for d in sorted(cv_dim.keys()):
        print(f"  {d}: CV = {cv_dim[d]:.4f}" if cv_dim[d] else f"  {d}: sem dados T1")

    # Fallback CV_ref para dimensões sem T1 (V6, V7, V8):
    # usar mediana dos CVs de todas as dimensões com T1
    all_cvs = [v for v in cv_dim.values() if v is not None]
    if all_cvs:
        all_cvs.sort()
        cv_fallback = all_cvs[len(all_cvs) // 2]
    else:
        cv_fallback = 0.5  # prior conservador
    print(f"  CV fallback (mediana global): {cv_fallback:.4f}")

    # ── 2. Ler registros codificados ──
    if not os.path.exists(FP_COD):
        print(f"ERRO: Arquivo de codificacao nao encontrado: {FP_COD}")
        print("Execute 30_preparar_codificacao_quali.py primeiro e preencha a planilha.")
        return

    wb_cod = openpyxl.load_workbook(FP_COD, read_only=True)
    ws_cod = wb_cod["CODIFICACAO"]

    # Ler cabeçalhos
    headers = []
    for row in ws_cod.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)

    registros_conv = []
    n_sem_codificacao = 0
    n_convertidos = 0
    n_excluidos_conv = 0

    for row in ws_cod.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue

        rec = dict(zip(headers, row))

        # Filtrar estudos excluídos
        study_id = safe_int(rec.get("Study_ID"))
        if study_id in IDS_EXCLUIDOS:
            continue

        tier = rec.get("Tier", "")

        # T1_QUANTI já capturado como T1 na leitura de bd_extracao
        if tier == "T1_QUANTI":
            continue

        direction = safe_int(rec.get("Direcao_efeito"))
        intensity = safe_int(rec.get("Intensidade"))

        # Verificar se foi codificado
        if direction is None:
            n_sem_codificacao += 1
            continue

        # Extrair n codificados
        n_T_cod = safe_int(rec.get("n_T_codificado"))
        n_C_cod = safe_int(rec.get("n_C_codificado"))

        # Extrair n de colunas reportadas (fallback)
        n_reportados_str = str(rec.get("n_reportados", ""))
        ns_report = []
        if n_reportados_str:
            for token in re.findall(r"\d+", n_reportados_str):
                ns_report.append(int(token))

        # Determinar n_T e n_C
        n_T = n_T_cod
        n_C = n_C_cod
        if n_T is None and n_C is None and ns_report:
            if len(ns_report) >= 2:
                n_T, n_C = ns_report[0], ns_report[1]
            elif len(ns_report) == 1:
                n_T = n_C = ns_report[0] // 2
        if n_T is None:
            n_T = N_FALLBACK // 2
        if n_C is None:
            n_C = N_FALLBACK // 2

        dim = rec.get("Dimensao", "")
        cv_ref = cv_dim.get(dim)
        if cv_ref is None:
            cv_ref = cv_fallback  # V6, V7, V8 sem T1
        sigma_conv = 0.0

        lnrr = None
        vi = None

        if tier == "T2a":
            # Extrair melhor p-value
            pval_str = str(rec.get("p_values_extraidos", ""))
            pvals = []
            for token in re.findall(r"[\d.]+", pval_str):
                try:
                    p = float(token)
                    if 0 < p < 1:
                        pvals.append(p)
                except ValueError:
                    pass

            if pvals:
                # Usar o menor p-value (efeito mais forte reportado)
                p_best = min(pvals)
                lnrr, vi = converter_t2a(p_best, direction, n_T, n_C, cv_ref)
                # Inflação por incerteza de conversão
                sigma_conv = 0.15  # ~ 15% de incerteza na conversão p→d
                vi = vi * (1 + sigma_conv**2)
            else:
                # Sem p-value válido: fallback para ordinal
                if intensity is None:
                    intensity = 2
                lnrr, vi = converter_ordinal(direction, intensity, n_T, n_C)
                sigma_conv = 0.30
                vi = vi * (1 + sigma_conv**2)

        elif tier == "T2b":
            # TEM_STAT mas sem p-value explícito
            # Usar threshold conservador p = 0.05 se intensidade ≥ 2, p = 0.10 se fraca
            if intensity is None:
                intensity = 2
            p_threshold = 0.05 if intensity >= 2 else 0.10
            lnrr, vi = converter_t2a(p_threshold, direction, n_T, n_C, cv_ref)
            sigma_conv = 0.25
            vi = vi * (1 + sigma_conv**2)

        elif tier in ("T3", "T4"):
            if intensity is None:
                intensity = 2
            lnrr, vi = converter_ordinal(direction, intensity, n_T, n_C)
            sigma_conv = 0.35 if tier == "T3" else 0.50
            vi = vi * (1 + sigma_conv**2)
            if tier == "T4":
                vi = vi * INFLATE_T4

        if lnrr is not None and vi is not None:
            registros_conv.append({
                "Study_ID": rec.get("Study_ID"),
                "Study": rec.get("Study"),
                "DOI": rec.get("DOI"),
                "Year": rec.get("Year"),
                "Dimensao": dim,
                "Dimensao_Label": rec.get("Dimensao_Label"),
                "Proxy": rec.get("Proxy"),
                "n_T": n_T,
                "m_T": None,
                "sd_T": None,
                "n_C": n_C,
                "m_C": None,
                "sd_C": None,
                "Tipo_Intervencao": rec.get("Tipo_Intervencao", rec.get("Tipo_Intervencao")),
                "Regiao": rec.get("Regiao", rec.get("Regiao")),
                "Tipo_Comunidade": rec.get("Tipo_Comunidade", rec.get("Tipo_Comunidade")),
                "NOS": rec.get("NOS"),
                "Tier": tier,
                "lnRR": round(lnrr, 6),
                "vi": round(vi, 6),
                "sigma_conv": round(sigma_conv, 4),
                "Direcao_codificada": direction,
                "Intensidade_codificada": intensity,
                "Confianca": rec.get("Confianca_codificacao", ""),
            })
            n_convertidos += 1
        else:
            n_excluidos_conv += 1

    wb_cod.close()

    # ── 3. Combinar T1 + convertidos ──
    todos = registros_t1 + registros_conv

    # ── 4. Exportar ──
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "DADOS_CONVERTIDOS"

    out_headers = [
        "Study_ID", "Study", "DOI", "Year", "Dimensao", "Dimensao_Label",
        "Proxy", "n_T", "m_T", "sd_T", "n_C", "m_C", "sd_C",
        "Tipo_Intervencao", "Regiao", "Tipo_Comunidade", "NOS",
        "Tier", "lnRR", "vi", "sigma_conv",
    ]

    fill_h = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font_h = Font(bold=True, color="FFFFFF")
    fill_t1 = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_t2 = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    fill_t34 = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for j, h in enumerate(out_headers, 1):
        cell = ws_out.cell(row=1, column=j, value=h)
        cell.fill = fill_h
        cell.font = font_h

    for i, r in enumerate(todos, 2):
        for j, h in enumerate(out_headers, 1):
            cell = ws_out.cell(row=i, column=j, value=r.get(h))
            t = r.get("Tier", "")
            if t == "T1":
                cell.fill = fill_t1
            elif t.startswith("T2"):
                cell.fill = fill_t2
            else:
                cell.fill = fill_t34

    # Sheet resumo
    ws_res = wb_out.create_sheet("RESUMO")
    ws_res.cell(row=1, column=1, value="Métrica").font = Font(bold=True)
    ws_res.cell(row=1, column=2, value="Valor").font = Font(bold=True)
    resumo = [
        ("Total registros convertidos", len(todos)),
        ("T1 (quantitativo direto)", len(registros_t1)),
        ("T2a (p-value → lnRR)", sum(1 for r in registros_conv if r["Tier"] == "T2a")),
        ("T2b (ANOVA/KW → lnRR)", sum(1 for r in registros_conv if r["Tier"] == "T2b")),
        ("T3 (ordinal Results=SIM)", sum(1 for r in registros_conv if r["Tier"] == "T3")),
        ("T4 (qualitativo puro)", sum(1 for r in registros_conv if r["Tier"] == "T4")),
        ("Sem codificação (pendente)", n_sem_codificacao),
        ("Excluídos na conversão", n_excluidos_conv),
    ]
    for i, (k, v) in enumerate(resumo, 2):
        ws_res.cell(row=i, column=1, value=k)
        ws_res.cell(row=i, column=2, value=v)

    # Resumo por dimensão
    ws_res.cell(row=12, column=1, value="Dimensão").font = Font(bold=True)
    ws_res.cell(row=12, column=2, value="k total").font = Font(bold=True)
    ws_res.cell(row=12, column=3, value="k T1").font = Font(bold=True)
    ws_res.cell(row=12, column=4, value="k T2").font = Font(bold=True)
    ws_res.cell(row=12, column=5, value="k T3+T4").font = Font(bold=True)
    ws_res.cell(row=12, column=6, value="lnRR médio").font = Font(bold=True)

    dims = ["V1", "V2", "V3", "V4", "V5", "V6", "V7", "V8"]
    for idx, dim in enumerate(dims, 13):
        dim_recs = [r for r in todos if r["Dimensao"] == dim]
        ws_res.cell(row=idx, column=1, value=dim)
        ws_res.cell(row=idx, column=2, value=len(dim_recs))
        ws_res.cell(row=idx, column=3, value=sum(1 for r in dim_recs if r["Tier"] == "T1"))
        ws_res.cell(row=idx, column=4, value=sum(1 for r in dim_recs if r["Tier"].startswith("T2")))
        ws_res.cell(row=idx, column=5, value=sum(1 for r in dim_recs if r["Tier"] in ("T3", "T4")))
        if dim_recs:
            lnrrs = [r["lnRR"] for r in dim_recs if r["lnRR"] is not None]
            ws_res.cell(row=idx, column=6, value=round(sum(lnrrs) / len(lnrrs), 4) if lnrrs else None)

    wb_out.save(FP_OUT_XLSX)

    # CSV backup
    import csv
    with open(FP_OUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=out_headers)
        writer.writeheader()
        for r in todos:
            writer.writerow({h: r.get(h) for h in out_headers})

    print(f"\n[OK] Banco convertido salvo em: {FP_OUT_XLSX}")
    print(f"[OK] CSV backup salvo em: {FP_OUT_CSV}")
    print(f"\n=== RESUMO DA CONVERSAO ===")
    for k, v in resumo:
        print(f"  {k}: {v}")

    print(f"\n=== POR DIMENSAO ===")
    for dim in dims:
        dim_recs = [r for r in todos if r["Dimensao"] == dim]
        k_total = len(dim_recs)
        k_t1 = sum(1 for r in dim_recs if r["Tier"] == "T1")
        k_t2 = sum(1 for r in dim_recs if r["Tier"].startswith("T2"))
        k_t34 = sum(1 for r in dim_recs if r["Tier"] in ("T3", "T4"))
        status = "Confirmatoria" if k_total >= 10 else ("Exploratoria" if k_total >= 3 else "Narrativa")
        print(f"  {dim}: k={k_total} (T1={k_t1}, T2={k_t2}, T3/T4={k_t34}) -> {status}")


if __name__ == "__main__":
    main()
