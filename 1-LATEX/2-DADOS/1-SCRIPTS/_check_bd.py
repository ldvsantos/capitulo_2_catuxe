# Quick check bd_extracao original
import openpyxl

wb = openpyxl.load_workbook(
    r"c:\Users\vidal\OneDrive\Documentos\13 - CLONEGIT"
    r"\artigo_2_catuxe\1-LATEX\2-DADOS\2-BANCO_DADOS"
    r"\2-DADOS_TABULADOS\bd_extracao.xlsx"
)
ws = wb.active
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

print("=== BD_EXTRACAO ORIGINAL - COLUNAS 9-21 ===")
for c in range(9, 22):
    cn = headers[c - 1]
    filled = 0
    for r in range(2, 254):
        v = ws.cell(r, c).value
        if v is not None and str(v).strip():
            filled += 1
    print(f"  {cn:<25}: {filled}/252")

print("\n=== ROW 2 COMPLETA ===")
for c in range(1, 22):
    v = ws.cell(2, c).value
    print(f"  {headers[c-1]:<20}: {repr(v)[:80]}")

wb.close()
