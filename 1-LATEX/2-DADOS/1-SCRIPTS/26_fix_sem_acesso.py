# -*- coding: utf-8 -*-
"""26_fix_sem_acesso.py — Marcar corretamente os 9 estudos sem PDF."""
import sys, shutil, tempfile
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill

sys.stdout.reconfigure(encoding="utf-8")

TAB = Path(r"c:\Users\vidal\OneDrive\Documentos\13 - CLONEGIT"
           r"\artigo_2_catuxe\1-LATEX\2-DADOS\2-BANCO_DADOS\2-DADOS_TABULADOS")
BD_PATH = TAB / "bd_extracao_PREENCHIDO.xlsx"

IDS_SEM_ACESSO = {1, 23, 24, 26, 28, 33, 34, 40, 42}
FILL_RED = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

tmp = Path(tempfile.gettempdir()) / "fix_sem.xlsx"
shutil.copy2(BD_PATH, tmp)

wb = openpyxl.load_workbook(tmp)
ws = wb.active
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
COL = {h: i + 1 for i, h in enumerate(headers)}

count = 0
for r in range(2, ws.max_row + 1):
    sid = ws.cell(r, COL["Study_ID"]).value
    if sid in IDS_SEM_ACESSO:
        old_notas = str(ws.cell(r, COL["Notas"]).value or "")
        if "[SEM ACESSO]" not in old_notas:
            new_notas = f"[SEM ACESSO] {old_notas}".strip()
            ws.cell(r, COL["Notas"]).value = new_notas
            ws.cell(r, COL["Notas"]).fill = FILL_RED
            count += 1

wb.save(tmp)
wb.close()
shutil.copy2(tmp, BD_PATH)

print(f"Atualizadas {count} celulas Notas com [SEM ACESSO]")
print(f"IDs marcados: {sorted(IDS_SEM_ACESSO)}")
