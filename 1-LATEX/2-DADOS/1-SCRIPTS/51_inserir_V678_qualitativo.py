#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
51_inserir_V678_qualitativo.py
==============================
Insere as 144 linhas V6/V7/V8 (codificadas) no arquivo
bd_codificacao_qualitativa_V8.xlsx, lendo os dados do
bd_extracao_PREENCHIDO_V8.xlsx.
"""
import os
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
DIR_DADOS = os.path.join(BASE, "..", "2-BANCO_DADOS", "2-DADOS_TABULADOS")
FP_PRENCH = os.path.join(DIR_DADOS, "bd_extracao_PREENCHIDO_V8.xlsx")
FP_QUAL = os.path.join(DIR_DADOS, "bd_codificacao_qualitativa_V8.xlsx")

LABEL_MAP = {
    "V6": "Vitalidade Linguistica",
    "V7": "Integracao ao Mercado",
    "V8": "Exposicao Climatica",
}

# ── Read PREENCHIDO V6-V8 rows ───────────────────────────────────
wb_p = openpyxl.load_workbook(FP_PRENCH, read_only=True)
ws_p = wb_p.active
hp = [c.value for c in ws_p[1]]
idx_p = {h: i for i, h in enumerate(hp)}

v678_rows = []
for row in ws_p.iter_rows(min_row=2, values_only=True):
    dim = row[idx_p["Dimensao"]]
    if dim not in ("V6", "V7", "V8"):
        continue
    v678_rows.append({
        "Study_ID": row[idx_p["Study_ID"]],
        "Study": row[idx_p["Study"]],
        "DOI": row[idx_p["DOI"]],
        "Year": row[idx_p["Year"]],
        "Dimensao": row[idx_p["Dimensao"]],
        "Dimensao_Label": row[idx_p["Dimensao_Label"]],
        "Proxy": row[idx_p.get("Proxy", 0)],
        "Tier": row[idx_p["Tier"]],
        "p_values_extraidos": None,
        "n_reportados": None,
        "Comparacoes_extraidas": None,
        "Notas": row[idx_p["Notas"]],
        "Direcao_efeito": row[idx_p["Direcao_efeito"]],
        "Intensidade": row[idx_p["Intensidade"]],
        "n_T_codificado": None,
        "n_C_codificado": None,
        "Revisor": row[idx_p["Revisor"]],
        "Confianca_codificacao": row[idx_p["Confianca_codificacao"]],
        "Notas_codificador": row[idx_p["Notas_codificador"]],
    })
wb_p.close()

print(f"V6-V8 rows read from PREENCHIDO: {len(v678_rows)}")

# ── Insert into CODIFICACAO ───────────────────────────────────────
wb_q = openpyxl.load_workbook(FP_QUAL)
ws_q = wb_q["CODIFICACAO"]
hq = [c.value for c in ws_q[1]]

next_row = ws_q.max_row + 1
inserted = 0

for rec in v678_rows:
    for j, h in enumerate(hq, 1):
        ws_q.cell(row=next_row, column=j, value=rec.get(h))
    next_row += 1
    inserted += 1

wb_q.save(FP_QUAL)
print(f"Inserted {inserted} rows into CODIFICACAO (total now: {ws_q.max_row - 1})")

# Verify
from collections import Counter
dim_col = hq.index("Dimensao")
dir_col = hq.index("Direcao_efeito")
for dim in ["V6", "V7", "V8"]:
    dirs = []
    for r in ws_q.iter_rows(min_row=2, values_only=True):
        if r[dim_col] == dim:
            dirs.append(r[dir_col])
    c = Counter(dirs)
    print(f"  {dim}: {len(dirs)} rows, Dir={dict(c)}")
