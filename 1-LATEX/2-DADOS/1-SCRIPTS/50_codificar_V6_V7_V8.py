#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
50_codificar_V6_V7_V8.py
=========================
Codifica as dimensões V6 (Vitalidade Linguística), V7 (Integração ao Mercado)
e V8 (Exposição Climática) para os 48 estudos no banco bd_extracao_PREENCHIDO_V8.xlsx.

Decisões foram tomadas com base em:
  - Título do artigo
  - Notas extraídas automaticamente (keywords, p-values, tipo de análise)
  - Abstract (quando [SEM ACESSO] mas abstract disponível nas notas)
  - Região e tipo de comunidade
  - Codificação prévia de V1 para contexto

Critérios seguem o PROTOCOLO_CODIFICACAO_REVISORES.md Seção 3.

Author: Diego Vidal | 2026-03-06
"""

import os
import openpyxl
from openpyxl.styles import PatternFill

# ── Paths ─────────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
DIR_DADOS = os.path.join(BASE, "..", "2-BANCO_DADOS", "2-DADOS_TABULADOS")
FP_IN = os.path.join(DIR_DADOS, "bd_extracao_PREENCHIDO_V8.xlsx")
FP_OUT = os.path.join(DIR_DADOS, "bd_extracao_PREENCHIDO_V8.xlsx")  # overwrite

# ── Default coding for dimensions not addressed by study ──────────
DEFAULT = {
    "dir": 0,
    "int": 1,
    "tier": "T4",
    "conf": "baixa",
    "notes": "Estudo não aborda esta dimensão de forma direta.",
    "revisor": "DV-AI",
}

# ══════════════════════════════════════════════════════════════════
#  CODING MATRIX — Override dictionary
#  Key: (Study_ID, Dimensao) → dict with dir, int, tier, conf, notes
# ══════════════════════════════════════════════════════════════════

OVERRIDES = {
    # ────────────────────────── V6 (Vitalidade Linguística) ──────
    # ALL studies → default (Dir=0): nenhum estudo avalia explicitamente
    # a vitalidade linguística. Estudos etnobotânicos que listam nomes
    # vernaculares sem avaliar se estão sendo perdidos → Dir=0, Conf=baixa
    # (conforme protocolo §3, nota de V6).
    # Nenhum override necessário para V6.

    # ────────────────────────── V7 (Integração ao Mercado) ────────

    # Study 4: Romero-Silva — "commodification of the commons" em landraces no Andes
    (4, "V7"): {
        "dir": 1, "int": 2, "tier": "T3", "conf": "moderada",
        "notes": ("Título indica 'commodification of the commons': pressão de mercado "
                  "sobre variedades tradicionais de feijão andino. Estudo compara "
                  "morfologia de landraces frente à substituição por cultivares comerciais."),
    },
    # Study 11: Andreotti — quinoa smallholder organizations, Peru
    (11, "V7"): {
        "dir": 0, "int": 1, "tier": "T3", "conf": "moderada",
        "notes": ("Quinoa é commodity global; estudo explora opções de organização de "
                  "smallholders para manejo de agrobiodiversidade. Direção do efeito de "
                  "mercado sobre agrobiodiversidade é ambígua (boom pode simplificar "
                  "ou diversificar). Codificado como neutro por incerteza direcional."),
    },
    # Study 48: Rodríguez/Ørting — trends of on-farm conservation of ahipa, Bolivia
    (48, "V7"): {
        "dir": 1, "int": 2, "tier": "T3", "conf": "moderada",
        "notes": ("Conservação on-farm de ahipa (Pachyrhizus ahipa) em declínio ao longo "
                  "de 18 anos (1994-2012). Substituição por cultivos comerciais é driver "
                  "provável conforme título ('trends and drivers'). p<0.001 ANOVA para "
                  "diferenças entre sítios e períodos."),
    },

    # ────────────────────────── V8 (Exposição Climática) ──────────

    # Study 15: Tran — erosion of TEK under hydrosocial rupture, Mekong
    (15, "V8"): {
        "dir": 1, "int": 2, "tier": "T4", "conf": "moderada",
        "notes": ("'Hydrosocial rupture' no Mekong: alteração hidrológica (barragens + "
                  "mudanças climáticas) causa erosão de TEK. Título indica impacto negativo "
                  "sobre práticas tradicionais das comunidades de planície de inundação."),
    },
    # Study 22: Aniah — climate variability adaptation, Ghana savanna
    (22, "V8"): {
        "dir": 1, "int": 2, "tier": "T2b", "conf": "moderada",
        "notes": ("'Livelihood adaptation to climate variability and ecological changes' "
                  "na zona agro-ecológica de savana de Ghana. Estudo com ANOVA/KW confirma "
                  "variabilidade climática como estressor sobre meios de vida de smallholders. "
                  "Adaptação implica que o sistema biocultural está sob pressão."),
    },
    # Study 25: Alemayehu — farmers' traditional knowledge for climate forecast, Ethiopia
    (25, "V8"): {
        "dir": -1, "int": 2, "tier": "T4", "conf": "moderada",
        "notes": ("'Farmers traditional knowledge on climate change and weather forecast': "
                  "TEK funciona como estratégia de adaptação climática. Agricultores usam "
                  "indicadores tradicionais (bioindicadores, fenologia) para previsão "
                  "climática → redução de vulnerabilidade do sistema biocultural."),
    },
    # Study 26: Zeleke — vulnerability assessment to climate change, Ethiopia
    (26, "V8"): {
        "dir": 1, "int": 2, "tier": "T4", "conf": "moderada",
        "notes": ("[SEM ACESSO] Abstract: 'quantifies vulnerability using indicator-based "
                  "framework of 15 indicators—six for exposure, four for sensitivity, five "
                  "for adaptive capacity'. Avaliação de vulnerabilidade climática de "
                  "smallholders na bacia do Awash. Direção +1 inferida do framework: clima "
                  "como driver de vulnerabilidade."),
    },
    # Study 35: Mondal — traditional knowledge to forecast flood, Bangladesh
    (35, "V8"): {
        "dir": -1, "int": 2, "tier": "T2a", "conf": "moderada",
        "notes": ("'Use of traditional knowledge to forecast flood': TEK para previsão de "
                  "enchentes em planície de inundação de Bangladesh. p<.001 para acurácia "
                  "de previsão tradicional. TEK como estratégia de adaptação funcional → "
                  "redução de vulnerabilidade."),
    },
    # Study 37: Arce — potato seed networks, seasons with/without acute stress
    (37, "V8"): {
        "dir": 1, "int": 2, "tier": "T3", "conf": "moderada",
        "notes": ("'Seasons With and Without Acute Stress' nos Andes: estresse climático "
                  "(seca/geada) afeta redes de sementes de batata. Comparação de provisões "
                  "e aquisições com stress (n=187/568) vs sem stress (n=582/357) indica "
                  "perturbação do sistema de agrobiodiversidade por eventos extremos."),
    },
    # Study 40: Avila — adaptive management facing extreme climate events, Amazonia
    (40, "V8"): {
        "dir": 1, "int": 2, "tier": "T4", "conf": "moderada",
        "notes": ("[SEM ACESSO] 'Adaptive Management Strategies of Local Communities in "
                  "Amazonian Floodplain Ecosystems in the Face of Extreme Climate Events'. "
                  "Eventos climáticos extremos afetando segurança alimentar e sistemas de "
                  "manejo na bacia do Solimões. Upgrade intensidade de 1→2 dado escopo "
                  "explícito do título."),
    },
    # Study 41: Mugi-Ngenga — socio-economic factors and climate adaptation, Kenya
    (41, "V8"): {
        "dir": 1, "int": 2, "tier": "T2b", "conf": "moderada",
        "notes": ("'Household socio-economic factors influencing level of adaptation to "
                  "climate variability' nas zonas secas do Leste do Quênia. Regressão "
                  "logística para determinantes de adaptação. Variabilidade climática como "
                  "estressor requerendo adaptação → vulnerability do sistema."),
    },
    # Study 47: Rodríguez-Cruz — disaster context (Hurricane Maria), Puerto Rico
    (47, "V8"): {
        "dir": 1, "int": 3, "tier": "T2a", "conf": "alta",
        "notes": ("'Social-ecological interactions in a disaster context: Puerto Rican "
                  "farmer households food security'. Furacão Maria (2017) como evento "
                  "climático extremo com impacto severo. n=[401, 154, 124, 123], p-values "
                  "[0.040, 0.030, 0.034, 0.017]. Intensidade 3: catástrofe climática "
                  "documentada com evidência estatística robusta."),
    },
}


def apply_coding(fp_in, fp_out):
    """Open workbook, apply coding to V6/V7/V8 rows, save."""
    wb = openpyxl.load_workbook(fp_in)
    ws = wb.active

    hdrs = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(hdrs)}

    col_sid = idx["Study_ID"]
    col_dim = idx["Dimensao"]
    col_tier = idx["Tier"]
    col_dir = idx["Direcao_efeito"]
    col_int = idx["Intensidade"]
    col_conf = idx["Confianca_codificacao"]
    col_notes_cod = idx["Notas_codificador"]
    col_revisor = idx["Revisor"]

    # highlight fill for newly coded cells
    fill_coded = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    stats = {"updated": 0, "overridden": 0, "default": 0}

    for row_idx in range(2, ws.max_row + 1):
        dim = ws.cell(row=row_idx, column=col_dim + 1).value
        if dim not in ("V6", "V7", "V8"):
            continue

        sid = ws.cell(row=row_idx, column=col_sid + 1).value
        key = (sid, dim)

        if key in OVERRIDES:
            coding = OVERRIDES[key]
            stats["overridden"] += 1
        else:
            coding = DEFAULT
            stats["default"] += 1

        # Apply coding
        ws.cell(row=row_idx, column=col_dir + 1).value = coding.get("dir", DEFAULT["dir"])
        ws.cell(row=row_idx, column=col_int + 1).value = coding.get("int", DEFAULT["int"])
        ws.cell(row=row_idx, column=col_tier + 1).value = coding.get("tier", DEFAULT["tier"])
        ws.cell(row=row_idx, column=col_conf + 1).value = coding.get("conf", DEFAULT["conf"])
        ws.cell(row=row_idx, column=col_notes_cod + 1).value = coding.get("notes", DEFAULT["notes"])
        ws.cell(row=row_idx, column=col_revisor + 1).value = coding.get("revisor", DEFAULT["revisor"])

        # Highlight
        for col in [col_dir, col_int, col_tier, col_conf, col_notes_cod, col_revisor]:
            ws.cell(row=row_idx, column=col + 1).fill = fill_coded

        stats["updated"] += 1

    wb.save(fp_out)
    print(f"Saved: {fp_out}")
    print(f"Total V6/V7/V8 rows updated: {stats['updated']}")
    print(f"  - With override (non-default): {stats['overridden']}")
    print(f"  - With default (Dir=0): {stats['default']}")

    # ── Diagnostic summary ──
    print("\n=== CODING SUMMARY ===")
    for dim_name in ["V6", "V7", "V8"]:
        dir_counts = {}
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=col_dim + 1).value == dim_name:
                d = ws.cell(row=row_idx, column=col_dir + 1).value
                dir_counts[d] = dir_counts.get(d, 0) + 1
        print(f"\n{dim_name}:")
        for d_val in sorted(dir_counts.keys(), key=lambda x: (x is None, x)):
            print(f"  Dir={d_val:>3}: {dir_counts[d_val]} studies")

    # ── Print overrides for verification ──
    print("\n=== OVERRIDES APPLIED ===")
    for key in sorted(OVERRIDES.keys()):
        sid, dim = key
        ov = OVERRIDES[key]
        print(f"  Study {sid:>2} {dim}: Dir={ov['dir']:>2}, Int={ov['int']}, "
              f"Tier={ov['tier']}, Conf={ov['conf']}")


if __name__ == "__main__":
    apply_coding(FP_IN, FP_OUT)
