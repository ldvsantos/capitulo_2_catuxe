# -*- coding: utf-8 -*-
"""
21_atualizar_preenchido.py
==========================
Copia dados do bd_extracao_V2.xlsx para bd_extracao_PREENCHIDO.xlsx,
preenchendo apenas células vazias (não sobrescreve dados existentes).
"""
import openpyxl, shutil, tempfile, sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

TAB = Path(r"c:\Users\vidal\OneDrive\Documentos\13 - CLONEGIT"
           r"\artigo_2_catuxe\1-LATEX\2-DADOS\2-BANCO_DADOS\2-DADOS_TABULADOS")

# Copiar V2 para temp (read-only source)
tmp_v = Path(tempfile.gettempdir()) / "v2_src.xlsx"
shutil.copy2(TAB / "bd_extracao_V2.xlsx", tmp_v)

# Ler V2 inteiro
wb_v = openpyxl.load_workbook(tmp_v, read_only=True)
ws_v = wb_v.active
hdr_v = [c.value for c in next(ws_v.iter_rows(min_row=1, max_row=1))]
rows_v = [list(r) for r in ws_v.iter_rows(min_row=2, values_only=True)]
wb_v.close()

# Build V2 lookup: (study_id, dimensao) -> dict of col_name -> value
dim_col_v = hdr_v.index("Dimensao")
v2_lookup = {}
for r in rows_v:
    key = (r[0], r[dim_col_v])
    v2_lookup[key] = {hdr_v[i]: r[i] for i in range(len(hdr_v)) if i < len(r)}

print(f"V2 carregado: {len(rows_v)} linhas, {len(v2_lookup)} chaves unicas")

# Abrir PREENCHIDO para edição (não read_only)
# Copiar para temp primeiro, editar, depois copiar de volta
tmp_p = Path(tempfile.gettempdir()) / "pre_edit.xlsx"
shutil.copy2(TAB / "bd_extracao_PREENCHIDO.xlsx", tmp_p)

wb_p = openpyxl.load_workbook(tmp_p)
ws_p = wb_p.active
hdr_p = [c.value for c in ws_p[1]]

dim_col_p = hdr_p.index("Dimensao")

# Colunas de dados a atualizar (skip metadata: Study_ID, Study, DOI, Title, Journal, Year, Dimensao, Dimensao_Label)
COLS_UPDATE = ["Proxy", "n_T", "m_T", "sd_T", "n_C", "m_C", "sd_C",
               "Tipo_Intervencao", "Regiao", "Tipo_Comunidade",
               "Tempo_Intervencao", "NOS", "Notas"]

# Map col name -> col index in PREENCHIDO
col_idx_p = {name: i for i, name in enumerate(hdr_p)}

updates = 0
updates_by_col = {c: 0 for c in COLS_UPDATE}

for row_num in range(2, ws_p.max_row + 1):
    sid = ws_p.cell(row=row_num, column=1).value
    dim = ws_p.cell(row=row_num, column=dim_col_p + 1).value
    key = (sid, dim)
    
    v2_data = v2_lookup.get(key)
    if v2_data is None:
        continue
    
    for col_name in COLS_UPDATE:
        if col_name not in col_idx_p or col_name not in v2_data:
            continue
        
        ci = col_idx_p[col_name] + 1  # openpyxl is 1-indexed
        current_val = ws_p.cell(row=row_num, column=ci).value
        new_val = v2_data[col_name]
        
        # Só preencher se vazio no PREENCHIDO e preenchido no V2
        is_empty = current_val is None or str(current_val).strip() == ""
        has_new = new_val is not None and str(new_val).strip() != ""
        
        if is_empty and has_new:
            ws_p.cell(row=row_num, column=ci, value=new_val)
            updates += 1
            updates_by_col[col_name] += 1

print(f"\nTotal de celulas atualizadas: {updates}")
print("\nAtualizacoes por coluna:")
for col, n in updates_by_col.items():
    if n > 0:
        print(f"  {col:<20}: +{n}")

# Salvar
wb_p.save(tmp_p)
wb_p.close()

# Copiar de volta para o destino
dst = TAB / "bd_extracao_PREENCHIDO.xlsx"
shutil.copy2(tmp_p, dst)
print(f"\nArquivo salvo: {dst}")

# Verificação final
wb_check = openpyxl.load_workbook(tmp_p, read_only=True)
ws_check = wb_check.active
hdr_check = [c.value for c in next(ws_check.iter_rows(min_row=1, max_row=1))]
rows_check = list(ws_check.iter_rows(min_row=2, values_only=True))
wb_check.close()

print("\n" + "=" * 80)
print("COBERTURA FINAL (bd_extracao_PREENCHIDO.xlsx ATUALIZADO)")
print("=" * 80)

for i, col_name in enumerate(hdr_check):
    n = sum(1 for r in rows_check if i < len(r) and r[i] is not None and str(r[i]).strip() != "")
    pct = n / len(rows_check) * 100
    print(f"  {col_name:<20}: {n:3d}/{len(rows_check)}  ({pct:.0f}%)")
