# -*- coding: utf-8 -*-
"""
23_extrair_novos_e_atualizar.py
================================
Pipeline completo:
  1) Extrai texto dos 13 novos PDFs
  2) Busca mean±SD, sample sizes, p-values
  3) Atualiza bd_extracao_PREENCHIDO.xlsx com dados novos
  4) Marca 9 estudos sem PDF como "Sem acesso"
"""
import os, re, json, sys, shutil, tempfile
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill

try:
    import fitz  # PyMuPDF
except ImportError:
    sys.exit("pip install PyMuPDF")

sys.stdout.reconfigure(encoding="utf-8")

ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def sanitize_for_excel(s):
    """Remove illegal XML characters that openpyxl rejects."""
    if not isinstance(s, str):
        return s
    return ILLEGAL_CHARS_RE.sub('', s)

# ── Paths ──
TAB = Path(r"c:\Users\vidal\OneDrive\Documentos\13 - CLONEGIT"
           r"\artigo_2_catuxe\1-LATEX\2-DADOS\2-BANCO_DADOS\2-DADOS_TABULADOS")
PDF_DIR = Path(r"c:\Users\vidal\OneDrive\Documentos\13 - CLONEGIT"
               r"\artigo_2_catuxe\1-LATEX\2-DADOS\2-BANCO_DADOS\1-ARTIGOS_SELECIONADOS")
BD_PATH = TAB / "bd_extracao_PREENCHIDO.xlsx"

HIGHLIGHT = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
HIGHLIGHT_GREEN = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
HIGHLIGHT_RED = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

# ── Mapeamento dos 13 novos PDFs → Study_ID ──
NEW_PDF_MAP = {
    "calvet-mir2015": 2,
    "mrd.2024.00024": 10,
    "1-s2.0-S2665972725003484": 12,
    "1-s2.0-S2405844022010933": 17,
    "Land Degrad Dev - 2020 - Nord": 18,
    "Land Degrad Dev - 2022 - Tabe": 21,
    "1-s2.0-S2405844019311880": 22,
    "Farmers-traditional-knowledge": 25,
    "Conservation Letters - 2017 - Fern": 29,
    "frascaroli2015": 31,
    "0350-75992500003M": 35,
    "10.1002@csc2.20620": 38,
    "mugi-ngenga2016": 41,
}

# IDs sem PDF (sem permissão de download)
IDS_SEM_ACESSO = {1, 23, 24, 26, 28, 33, 34, 40, 42}

# ── Regex patterns ──
RE_MEAN_SD = re.compile(
    r'(?<!\d{3}[-/])(?<!doi[.:])(?<!org/)(?<![A-Za-z])'
    r'(\d{1,6}\.?\d{0,4})\s*[±]\s*(\d{1,6}\.?\d{0,4})'
    r'(?![-/]\d)',
    re.UNICODE
)
RE_MEAN_PAREN = re.compile(
    r'(?:mean|average|M)\s*[=:]\s*(\d+\.?\d*)\s*[\(,]\s*(?:SD|SE|s\.d\.)\s*[=:]\s*(\d+\.?\d*)',
    re.IGNORECASE
)
RE_P_VALUE = re.compile(
    r'[Pp]\s*[<>=]\s*0?\.\d+|[Pp]\s*=\s*0?\.\d+',
)
RE_N_SAMPLE = re.compile(
    r'[Nn]\s*=\s*(\d+)',
)
RE_R_SQUARED = re.compile(
    r'[Rr]²?\s*=\s*0?\.\d+|R-squared\s*=\s*0?\.\d+',
    re.IGNORECASE
)
RE_CHI2 = re.compile(
    r'[Xχ]²\s*=\s*\d+\.?\d*|chi[- ]?square',
    re.IGNORECASE
)
RE_ANOVA = re.compile(
    r'ANOVA|[Ff]\s*[\(=]\s*\d|Kruskal|Mann[- ]Whitney|Wilcoxon',
    re.IGNORECASE
)
RE_REGRESSION = re.compile(
    r'regression|logistic|[Oo]dds ratio|OR\s*=|Exp\s*\(B\)|coefficient|β|beta',
    re.IGNORECASE
)


def extract_pdf_text(pdf_path, max_pages=None):
    """Extrai texto completo de um PDF."""
    doc = fitz.open(str(pdf_path))
    pages = doc if max_pages is None else doc[:max_pages]
    text = ""
    for page in pages:
        text += page.get_text() + "\n"
    doc.close()
    return text


def find_results_section(text):
    """Tenta isolar a seção Results/Discussion."""
    patterns = [
        r'(?i)\n\s*(?:3\.?\s*)?results?\s*(?:and discussion)?\s*\n',
        r'(?i)\n\s*(?:4\.?\s*)?results?\s*\n',
        r'(?i)\n\s*findings\s*\n',
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            start = m.start()
            # Find end (next major section)
            end_m = re.search(r'(?i)\n\s*(?:(?:4|5|6)\.?\s*)?(?:conclusion|discussion|acknowledg|reference)', text[start+100:])
            end = start + 100 + end_m.start() if end_m else len(text)
            return text[start:end]
    return text  # fallback: entire text


def analyze_pdf(pdf_path, study_id):
    """Analisa um PDF e retorna dicionário com dados extraídos."""
    text = extract_pdf_text(pdf_path)
    results_text = find_results_section(text)
    
    result = {
        "study_id": study_id,
        "n_chars": len(text),
        "results_chars": len(results_text),
        "mean_sd_pairs": [],
        "p_values": [],
        "sample_sizes": [],
        "has_regression": bool(RE_REGRESSION.search(text)),
        "has_anova": bool(RE_ANOVA.search(text)),
        "has_chi2": bool(RE_CHI2.search(text)),
        "r_squared": [],
        "results_preview": results_text[:500],
        "classification": "QUALIT",
    }
    
    # Mean±SD
    for m in RE_MEAN_SD.finditer(results_text):
        mean_val = float(m.group(1))
        sd_val = float(m.group(2))
        ctx = results_text[max(0, m.start()-60):m.end()+60].replace('\n', ' ')
        result["mean_sd_pairs"].append({
            "mean": mean_val, "sd": sd_val,
            "match": m.group(), "context": ctx
        })
    
    # Also check full text for mean±SD (skip references section)
    # Try to cut text before references
    ref_start = re.search(r'(?i)\n\s*references\s*\n', text)
    search_text = text[:ref_start.start()] if ref_start else text
    for m in RE_MEAN_SD.finditer(search_text):
        mean_val = float(m.group(1))
        sd_val = float(m.group(2))
        ctx = search_text[max(0, m.start()-60):m.end()+60].replace('\n', ' ')
        # Skip if looks like a DOI, ISSN, page range
        if re.search(r'doi|issn|https?://|\d{4}-\d{4}', ctx, re.I):
            continue
        # Check not already captured
        if not any(p["match"] == m.group() and abs(p["mean"]-mean_val) < 0.001 for p in result["mean_sd_pairs"]):
            result["mean_sd_pairs"].append({
                "mean": mean_val, "sd": sd_val,
                "match": m.group(), "context": ctx,
                "source": "fulltext"
            })
    
    # p-values
    for m in RE_P_VALUE.finditer(results_text):
        result["p_values"].append(m.group())
    
    # Sample sizes
    for m in RE_N_SAMPLE.finditer(text):
        n = int(m.group(1))
        if 5 < n < 100000:
            result["sample_sizes"].append(n)
    
    # R²
    for m in RE_R_SQUARED.finditer(text):
        result["r_squared"].append(m.group())
    
    # Classify
    n_msd = len(result["mean_sd_pairs"])
    n_p = len(result["p_values"])
    if n_msd >= 2 and n_p >= 1:
        result["classification"] = "IDEAL"
    elif n_msd >= 1:
        result["classification"] = "TEM_MEAN"
    elif result["has_regression"] or result["has_anova"]:
        result["classification"] = "TEM_STAT"
    elif n_p > 0:
        result["classification"] = "TEM_P"
    elif result["sample_sizes"]:
        result["classification"] = "TEM_N"
    
    return result


def build_notas(analysis):
    """Constrói string de notas a partir da análise."""
    parts = [f"[{analysis['classification']}]"]
    
    n_msd = len(analysis["mean_sd_pairs"])
    n_p = len(analysis["p_values"])
    n_s = len(set(analysis["sample_sizes"]))
    
    if n_msd:
        parts.append(f"{n_msd} mean±SD pairs")
        # Show first 3 (truncated, no DOIs)
        for pair in analysis["mean_sd_pairs"][:3]:
            ctx = pair['context'][:60].replace('\n', ' ')
            # skip contexts that look like references/DOIs
            if re.search(r'doi|https?://|issn', ctx, re.I):
                continue
            parts.append(f"  {pair['match']}: {ctx}")
    
    if n_p:
        unique_p = list(set(analysis["p_values"]))[:5]
        parts.append(f"p-values: {', '.join(unique_p)}")
    
    if n_s:
        sizes = sorted(set(analysis["sample_sizes"]))[:5]
        parts.append(f"sample sizes: {sizes}")
    
    if analysis["has_regression"]:
        parts.append("Has regression/logistic")
    if analysis["has_anova"]:
        parts.append("Has ANOVA/KW")
    if analysis["r_squared"]:
        parts.append(f"R²: {analysis['r_squared'][:3]}")
    
    return " | ".join(parts)


def build_proxy_from_analysis(analysis, dim):
    """Sugere proxy baseado no conteúdo do PDF."""
    text = analysis["results_preview"].lower()
    
    proxy_hints = {
        "V1": [("species richness", "Species richness"),
               ("diversity index", "Biodiversity index (Shannon/Simpson)"),
               ("cultivar", "Cultivar diversity"),
               ("crop diversity", "Crop species diversity"),
               ("plant species", "Plant species count"),
               ("agrobiodiversity", "Agrobiodiversity index")],
        "V2": [("knowledge", "Traditional knowledge score"),
               ("intergenerational", "Intergenerational knowledge transfer"),
               ("transmission", "Knowledge transmission indicator"),
               ("education", "Knowledge by education level")],
        "V3": [("network", "Social network indicator"),
               ("exchange", "Seed/resource exchange frequency"),
               ("cooperation", "Community cooperation index"),
               ("social capital", "Social capital indicator")],
        "V4": [("food security", "Food security indicator"),
               ("income", "Household income indicator"),
               ("livelihood", "Livelihood diversity index"),
               ("production", "Agricultural production indicator"),
               ("yield", "Crop yield indicator")],
        "V5": [("vulnerability", "Vulnerability index"),
               ("resilience", "Resilience indicator"),
               ("adaptive capacity", "Adaptive capacity score"),
               ("conservation", "Conservation status indicator"),
               ("documentation", "Knowledge documentation status")],
        "V6": [("climate", "Climate change perception indicator"),
               ("land use", "Land use change indicator"),
               ("degradation", "Environmental degradation indicator"),
               ("external", "External pressure indicator"),
               ("threat", "Threat assessment indicator")],
    }
    
    if dim in proxy_hints:
        for keyword, proxy in proxy_hints[dim]:
            if keyword in text:
                return proxy
    
    # Fallback
    defaults = {
        "V1": "Agrobiodiversity index / species richness",
        "V2": "Intergenerational knowledge transfer indicator",
        "V3": "Cultural complexity / social network indicator",
        "V4": "Food security / dietary diversity indicator",
        "V5": "Documentation status / knowledge formalization",
        "V6": "External pressure / land use change indicator",
    }
    return defaults.get(dim, "")


def main():
    print("=" * 80)
    print("EXTRAÇÃO DOS 13 NOVOS PDFs + ATUALIZAÇÃO DO PREENCHIDO")
    print("=" * 80)
    
    # ── 1. Extrair e analisar cada novo PDF ──
    analyses = {}
    pdfs = list(PDF_DIR.glob("*.pdf"))
    
    for pdf in pdfs:
        fname = pdf.stem
        bid = None
        for prefix, pid in NEW_PDF_MAP.items():
            if fname.startswith(prefix):
                bid = pid
                break
        if bid is None:
            continue
        
        print(f"\nAnalisando ID={bid}: {fname[:60]}...")
        try:
            analysis = analyze_pdf(pdf, bid)
            analyses[bid] = analysis
            print(f"  Classificação: {analysis['classification']}")
            print(f"  mean±SD: {len(analysis['mean_sd_pairs'])} pares")
            print(f"  p-values: {len(analysis['p_values'])}")
            print(f"  Sample sizes: {sorted(set(analysis['sample_sizes']))[:5]}")
            if analysis["has_regression"]:
                print(f"  Tem regressão/logistic")
            if analysis["has_anova"]:
                print(f"  Tem ANOVA/KW")
        except Exception as e:
            print(f"  ERRO: {e}")
    
    print(f"\n\nTotal analisados: {len(analyses)}/13")
    
    # ── 2. Mostrar mean±SD encontrados para possível uso ──
    print(f"\n{'='*80}")
    print("DADOS QUANTITATIVOS ENCONTRADOS NOS NOVOS PDFs")
    print("=" * 80)
    
    for bid in sorted(analyses):
        a = analyses[bid]
        if a["mean_sd_pairs"]:
            print(f"\n--- ID={bid} ({a['classification']}) ---")
            for i, pair in enumerate(a["mean_sd_pairs"][:10]):
                src = pair.get("source", "results")
                print(f"  [{src}] {pair['match']:>15}  >>  {pair['context'][:100]}")
    
    # ── 3. Abrir PREENCHIDO e atualizar ──
    print(f"\n{'='*80}")
    print("ATUALIZANDO bd_extracao_PREENCHIDO.xlsx")
    print("=" * 80)
    
    # Copy to temp
    tmp = Path(tempfile.gettempdir()) / "pre_update.xlsx"
    shutil.copy2(BD_PATH, tmp)
    
    wb = openpyxl.load_workbook(tmp)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    COL = {h: i + 1 for i, h in enumerate(headers)}
    
    updates = {"notas": 0, "proxy": 0, "sem_acesso": 0}
    
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, COL["Study_ID"]).value
        dim = ws.cell(r, COL["Dimensao"]).value
        if not sid:
            continue
        
        # ── Marcar estudos sem acesso ──
        if sid in IDS_SEM_ACESSO:
            current_notas = ws.cell(r, COL["Notas"]).value or ""
            if "[SEM ACESSO]" not in current_notas:
                prefix = "[SEM ACESSO - PDF indisponivel] "
                ws.cell(r, COL["Notas"]).value = sanitize_for_excel(prefix + current_notas)
                ws.cell(r, COL["Notas"]).fill = HIGHLIGHT_RED
                updates["sem_acesso"] += 1
        
        # ── Atualizar com dados dos novos PDFs ──
        if sid in analyses:
            a = analyses[sid]
            
            # Atualizar Notas se vazia ou genérica
            current_notas = ws.cell(r, COL["Notas"]).value or ""
            if "[ABSTRACT]" in current_notas or not current_notas.strip():
                new_notas = sanitize_for_excel(build_notas(a))
                ws.cell(r, COL["Notas"]).value = new_notas
                ws.cell(r, COL["Notas"]).fill = HIGHLIGHT
                updates["notas"] += 1
            
            # Atualizar Proxy se genérico (DEFAULT_PROXY)
            current_proxy = ws.cell(r, COL["Proxy"]).value or ""
            generic_proxies = [
                "Agrobiodiversity index / species richness",
                "Intergenerational knowledge transfer indicator",
                "Cultural complexity / social network indicator",
                "Food security / dietary diversity indicator",
                "Documentation status / knowledge formalization",
                "External pressure / land use change indicator",
            ]
            if current_proxy in generic_proxies or not current_proxy.strip():
                new_proxy = build_proxy_from_analysis(a, dim)
                if new_proxy and new_proxy != current_proxy:
                    ws.cell(r, COL["Proxy"]).value = new_proxy
                    ws.cell(r, COL["Proxy"]).fill = HIGHLIGHT
                    updates["proxy"] += 1
    
    # ── 4. Inserir mean±SD para estudos com dados claros ──
    # Análise manual dos novos PDFs que têm mean±SD com grupos comparáveis
    
    # ID=21 Tabe-Ojong 2022 - Soil conservation Thailand
    # Tem regressão logística/multinomial, não mean±SD com T/C direto
    # Mas tem odds ratios e coefficients
    
    # ID=41 Mugi-Ngenga 2016 - Kenya smallholder
    # Tem regressão, provavelmente coeficientes e p-values
    
    # ID=18 Nord 2020 - Farmer perceptions
    # Tem ANOVA/KW
    
    # ID=22 Aniah 2019 - Climate adaptation Ghana  
    # Provavelmente qualit/mixed
    
    # Para cada novo estudo com mean±SD, inserir dados quando temos T/C claros
    # Isso requer revisão manual; por agora, atualizamos Notas e Proxy
    
    # Salvar
    wb.save(tmp)
    wb.close()
    
    # Copiar de volta
    shutil.copy2(tmp, BD_PATH)
    print(f"\nSalvo: {BD_PATH}")
    
    print(f"\nAtualizações:")
    for k, v in updates.items():
        print(f"  {k:<15}: {v}")
    
    # ── 5. Verificação final ──
    print(f"\n{'='*80}")
    print("COBERTURA FINAL")
    print("=" * 80)
    
    wb2 = openpyxl.load_workbook(tmp, read_only=True)
    ws2 = wb2.active
    total = ws2.max_row - 1
    
    for col_name in ["Proxy", "n_T", "m_T", "sd_T", "n_C", "m_C", "sd_C",
                     "Tipo_Intervencao", "Regiao", "Tipo_Comunidade", "Notas"]:
        ci = COL[col_name]
        filled = sum(1 for r in range(2, ws2.max_row + 1)
                     if ws2.cell(r, ci).value is not None and str(ws2.cell(r, ci).value).strip())
        print(f"  {col_name:<25}: {filled:3d}/{total} ({filled/total*100:.1f}%)")
    
    # Contagem excluindo sem acesso
    rows_com_acesso = total - len(IDS_SEM_ACESSO) * 6
    print(f"\n  Total linhas: {total}")
    print(f"  Estudos sem acesso a PDF: {len(IDS_SEM_ACESSO)} ({len(IDS_SEM_ACESSO)*6} linhas)")
    print(f"  Linhas com acesso: {rows_com_acesso}")
    
    # Contagem de classificação dos novos
    print(f"\n{'='*80}")
    print("CLASSIFICAÇÃO DOS 13 NOVOS PDFs")
    print("=" * 80)
    class_count = {}
    for bid in sorted(analyses):
        a = analyses[bid]
        c = a["classification"]
        class_count[c] = class_count.get(c, 0) + 1
        print(f"  ID={bid:3d}: {c:<10} | mean±SD={len(a['mean_sd_pairs']):2d} | p-val={len(a['p_values']):2d} | reg={'S' if a['has_regression'] else 'N'} | anova={'S' if a['has_anova'] else 'N'}")
    
    print(f"\nResumo classificação:")
    for c, n in sorted(class_count.items()):
        print(f"  {c}: {n}")
    
    wb2.close()
    
    # ── 6. Salvar extração JSON ──
    out_json = Path(tempfile.gettempdir()) / "extracao_novos_13.json"
    json_data = {}
    for bid, a in analyses.items():
        json_data[bid] = {
            "classification": a["classification"],
            "n_mean_sd": len(a["mean_sd_pairs"]),
            "n_p_values": len(a["p_values"]),
            "sample_sizes": sorted(set(a["sample_sizes"]))[:10],
            "has_regression": a["has_regression"],
            "has_anova": a["has_anova"],
            "mean_sd_top5": [
                {"mean": p["mean"], "sd": p["sd"], "ctx": p["context"][:100]}
                for p in a["mean_sd_pairs"][:5]
            ],
        }
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(json_data, f, indent=2, ensure_ascii=False)
    print(f"\nExtração salva: {out_json}")


if __name__ == "__main__":
    main()
