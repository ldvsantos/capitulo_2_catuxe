# -*- coding: utf-8 -*-
"""
18_preencher_completo.py
========================
Preenchimento AGRESSIVO do bd_extracao_PREENCHIDO.xlsx:
  - Regiao / Tipo_Comunidade / Tipo_Intervencao para TODOS 48 estudos
  - Proxy sugerido para cada dimensão
  - Notas com abstract para estudos sem PDF
  - mean±SD para estudos IDEAL (Sibanda, Rodriguez-Cruz)
  - mean±SD para estudos TEM_STAT onde possível
"""

import os, re, json
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill

try:
    import fitz
except ImportError:
    import sys; sys.exit("pip install PyMuPDF")

# ── Paths ────────────────────────────────────────────────────
BASE = Path(r"c:\Users\vidal\OneDrive\Documentos\13 - CLONEGIT"
            r"\artigo_2_catuxe\1-LATEX\2-DADOS\2-BANCO_DADOS")
PDF_DIR = r"\\?\ ".strip() + str(BASE / "1-ARTIGOS_SELECIONADOS")
BD_PREENCHIDO = BASE / "2-DADOS_TABULADOS" / "bd_extracao_PREENCHIDO.xlsx"
SEL42_PATH = BASE / "2-DADOS_TABULADOS" / "selecionados_42_completos.xlsx"
OUT_PATH = BASE / "2-DADOS_TABULADOS" / "bd_extracao_V2.xlsx"
EXTRACT_JSON = BASE / "3-OUTPUT" / "extracao_completa.json"

HIGHLIGHT = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
HIGHLIGHT_GREEN = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

# ── PDF → bd_id mapping ─────────────────────────────────────
PDF_TO_BDID = {
    "18-doyle": 36, "A-Biodiversity-Hotspot": 3, "Beyond-biodiversity": 39,
    "Biocultural-conservation-systems": 13, "Colourful-agrobiodiversity": 4,
    "Comparison-of-medicinal": 9, "Erosion-of-traditional": 15,
    "Ethnobotany-and-conservation": 16, "Ethnobotany-of-the-Aegadian": 27,
    "Exploring-farmers": 11, "Guardians-of-heritage": 20,
    "Interconnected-Nature": 19, "Mountain-Graticules": 6,
    "Stingless-bee-keeping": 7, "Strategies-for-managing": 5,
    "Towards-biocultural": 14, "Unearthing-Unevenness": 37,
    "Voices-Around-the-South": 30, "ajol-file-journals": 32,
    "The-Zo-perspective": 8,
    "1-s2.0-S1470160X20308037": 43, "s12231-018-9401-y": 44,
    "A-Quantitative-Study": 45, "1-s2.0-S1462901124001953": 46,
    "Socialecological": 47, "s10722-017-0544-y": 48,
}
EXCLUDE_PDFS = {"s13412-024-00888-3.pdf"}

# ── Region/Community/Intervention from title + abstract + affiliations ──
# Manual curation based on known metadata
STUDY_META = {
    # ID: Author (Year) - Title snippet
    1:  {"regiao": "America do Sul", "tipo_com": "Quilombola", "tipo_int": "Ethnobotanical survey"},          # Goncalves 2022 - Traditional Agriculture Quilombola
    2:  {"regiao": "Europa", "tipo_com": "Rural", "tipo_int": "Home garden knowledge transmission"},   # Calvet-Mir 2016 - Transmission Home Garden
    3:  {"regiao": "America do Sul", "tipo_com": "Agricultores", "tipo_int": "Agrobiodiversity assessment"}, # Bastos 2022 - Biodiversity Hotspot
    4:  {"regiao": "America do Sul", "tipo_com": "Agricultores", "tipo_int": "Agrobiodiversity assessment"}, # Romero-Silva 2026 - Bean landraces Chile
    5:  {"regiao": "America do Sul", "tipo_com": "Agricultores", "tipo_int": "Agrobiodiversity management"}, # de Sousa 2024 - Cerrado peasant farmers
    6:  {"regiao": "America do Sul", "tipo_com": "Rural", "tipo_int": "Landscape ethnoecology"},        # Sarmiento 2023 - Mountain Graticules
    7:  {"regiao": "America Central", "tipo_com": "Indigena", "tipo_int": "Traditional beekeeping"},     # Aldasoro Maya 2023 - Stingless bee Mexico
    8:  {"regiao": "America do Sul", "tipo_com": "Indigena", "tipo_int": "Biocultural landscape assessment"}, # Franco-Moraes 2023 - Zo'e perspective
    9:  {"regiao": "America do Sul", "tipo_com": "Rural", "tipo_int": "Medicinal plant survey"},        # Latorre 2018 - Medicinal plant rural vs urban
    10: {"regiao": "Africa", "tipo_com": "Rural", "tipo_int": "Agroforestry assessment"},               # Mobarak 2025 - Farm Trees Cultural Keystone
    11: {"regiao": "Europa", "tipo_com": "Agricultores", "tipo_int": "Agrobiodiversity management"},    # Andreotti 2023 - Farmers perspectives
    12: {"regiao": "Asia", "tipo_com": "Rural", "tipo_int": "Ethnobotanical survey"},                   # Suwardi 2025 - Ecological functions Indonesia
    13: {"regiao": "Europa", "tipo_com": "Rural", "tipo_int": "Biocultural conservation"},              # Plieninger 2023 - Mediterranean region
    14: {"regiao": "America do Sul", "tipo_com": "Indigena", "tipo_int": "Biocultural conservation"},   # Oloriz 2020 - Towards biocultural
    15: {"regiao": "Asia", "tipo_com": "Rural", "tipo_int": "Traditional knowledge erosion"},           # Tran 2025 - Erosion TEK hydrosocial
    16: {"regiao": "Oceania", "tipo_com": "Indigena", "tipo_int": "Ethnobotanical survey"},             # Saiba 2023 - Noken making Papua
    17: {"regiao": "America Central", "tipo_com": "Indigena", "tipo_int": "Agroforestry management"},   # Vazquez-Delfin 2022 - Traditional agroforestry Mexico
    18: {"regiao": "Europa", "tipo_com": "Agricultores", "tipo_int": "Farmer perception survey"},       # Nord 2020 - Farmer perceptions
    19: {"regiao": "Europa", "tipo_com": "Rural", "tipo_int": "Biosphere reserve assessment"},          # Rollo 2025 - Biosphere Reserves Portugal
    20: {"regiao": "Africa", "tipo_com": "Agricultores", "tipo_int": "Seed system assessment"},         # Sibanda 2025 - Women seed systems Zimbabwe
    21: {"regiao": "Africa", "tipo_com": "Agricultores", "tipo_int": "Soil conservation assessment"},   # Tabe-Ojong 2023 - Cassava smallholder
    22: {"regiao": "Africa", "tipo_com": "Agricultores", "tipo_int": "Climate adaptation assessment"}, # Aniah 2019 - Smallholder climate Ghana
    23: {"regiao": "Europa", "tipo_com": "Urbana", "tipo_int": "Nature-connectedness assessment"},     # Moller 2023 - Nature-Connectedness
    24: {"regiao": "America do Sul", "tipo_com": "Rural", "tipo_int": "Flora checklist"},               # Whaley 2019 - Vascular Flora Ica Peru
    25: {"regiao": "Africa", "tipo_com": "Agricultores", "tipo_int": "Traditional weather knowledge"}, # Alemayehu 2023 - Climate forecast Ethiopia
    26: {"regiao": "Africa", "tipo_com": "Agricultores", "tipo_int": "Vulnerability assessment"},      # Zeleke 2025 - Vulnerability Ethiopia
    27: {"regiao": "Europa", "tipo_com": "Rural", "tipo_int": "Ethnobotanical survey"},                 # La Rosa 2021 - Aegadian Islands Italy
    28: {"regiao": "America Central", "tipo_com": "Indigena", "tipo_int": "Biocultural diversity assessment"}, # Montoya-Greenheck 2018 - Ngobe Costa Rica
    29: {"regiao": "Global", "tipo_com": "Indigena", "tipo_int": "Indigenous storytelling review"},    # Fernandez-Llamazares 2018 - Storytelling
    30: {"regiao": "Europa", "tipo_com": "Rural", "tipo_int": "Herbal knowledge assessment"},          # Gerner 2025 - South Tyrolean herbal
    31: {"regiao": "Europa", "tipo_com": "Rural", "tipo_int": "Sacred grove conservation"},             # Frascaroli 2016 - Shrines Central Italy
    32: {"regiao": "Africa", "tipo_com": "Agricultores", "tipo_int": "Agroforestry assessment"},        # Edo 2024 - Homegarden Ethiopia
    33: {"regiao": "Global", "tipo_com": "Rural", "tipo_int": "Cultural species assessment"},          # Essien 2025 - Cultural Symbols wildlife
    34: {"regiao": "Asia", "tipo_com": "Rural", "tipo_int": "Bionic architecture assessment"},          # Haghverdi 2025 - Bionic architecture Iran
    35: {"regiao": "Asia", "tipo_com": "Rural", "tipo_int": "Traditional flood knowledge"},             # Mondal 2025 - Flood Bangladesh
    36: {"regiao": "Asia", "tipo_com": "Agricultores", "tipo_int": "Genetic resource management"},     # Doyle 2019 - Domestication aquaculture
    37: {"regiao": "America do Sul", "tipo_com": "Agricultores", "tipo_int": "Seed network analysis"},  # Arce 2018 - Potato seed Andes
    38: {"regiao": "America do Sul", "tipo_com": "Agricultores", "tipo_int": "On-farm conservation"},   # Garcia 2021 - Multicriteria CWR
    39: {"regiao": "Europa", "tipo_com": "Agricultores", "tipo_int": "Land sharing assessment"},        # Loos 2018 - Beyond Biodiversity Romania
    40: {"regiao": "America do Sul", "tipo_com": "Indigena", "tipo_int": "Adaptive management"},        # Avila 2021 - Amazonian Forests
    41: {"regiao": "Africa", "tipo_com": "Agricultores", "tipo_int": "Climate adaptation assessment"}, # Mugi-Ngenga 2016 - Kenya smallholder
    42: {"regiao": "America do Sul", "tipo_com": "Pescadores artesanais", "tipo_int": "Social-ecological monitoring"}, # Cantor 2025 - Dolphin fishers Brazil
    43: {"regiao": "Oceania", "tipo_com": "Indigena", "tipo_int": "Ecological knowledge indicator"},   # Aswani 2020 - Solomon Islands ITD
    44: {"regiao": "America do Sul", "tipo_com": "Indigena", "tipo_int": "Ethnobotanical assessment"}, # Bussmann 2018 - Chacobo Bolivia
    45: {"regiao": "Africa", "tipo_com": "Rural", "tipo_int": "Ethnobotanical survey"},                # Ghanimi 2022 - Wild edible Morocco
    46: {"regiao": "Africa", "tipo_com": "Indigena", "tipo_int": "Indigenous knowledge transmission"},  # Malapane 2024 - Vhavenda South Africa
    47: {"regiao": "Caribe", "tipo_com": "Agricultores", "tipo_int": "Post-disaster food security"},    # Rodriguez-Cruz 2022 - Puerto Rico
    48: {"regiao": "America do Sul", "tipo_com": "Agricultores", "tipo_int": "On-farm conservation"},   # Rodriguez 2018 - Ahipa Bolivia
}

# ── Quantitative data: ALL mean±SD for IDEAL and TEM_STAT studies ──
# Manually curated from extracao_completa.txt
QUANT_DATA = {
    # ──── ID=20 Sibanda (2025) Zimbabwe "Guardians of heritage" ────
    # Gender comparison: Women vs Men
    (20, "V2"): {
        "proxy": "Ethnobotanical knowledge score (gender)",
        "n_T": 100, "m_T": 38.96, "sd_T": 10.67,  # Women
        "n_C": 100, "m_C": 34.85, "sd_C": 11.65,   # Men
        "notas_quant": "t=-2.6* | Women mean=38.96±10.67 vs Men 34.85±11.65 | n=100 each | Sibanda Table 3",
    },
    # Age comparison: >=45y vs <45y
    (20, "V3"): {
        "proxy": "Ethnobotanical knowledge score (age >=45 vs <45)",
        "n_T": 100, "m_T": 41.38, "sd_T": 10.73,  # >=45y (traditional keepers)
        "n_C": 100, "m_C": 32.43, "sd_C": 10.14,   # <45y
        "notas_quant": "t=-6.062*** | >=45y 41.38±10.73 vs <45y 32.43±10.14 | n=100 each | Sibanda Table 3",
    },
    # Occupation: Agricultural vs Non-agricultural
    (20, "V4"): {
        "proxy": "Ethnobotanical knowledge score (occupation)",
        "n_T": 58, "m_T": 37.64, "sd_T": 11.44,   # Agricultural
        "n_C": 52, "m_C": 32.98, "sd_C": 11.62,    # Non-agricultural
        "notas_quant": "F=4.517* | Agricultural 37.64±11.44 n=58 vs Non-agri 32.98±11.62 n=52 | Also housewife 38.70±10.66 n=90",
    },
    # Education: No education vs University
    (20, "V5"): {
        "proxy": "Ethnobotanical knowledge score (education)",
        "n_T": 123, "m_T": 39.94, "sd_T": 10.59,  # No formal education
        "n_C": 8, "m_C": 21.00, "sd_C": 5.24,      # University
        "notas_quant": "F=11.967*** | No formal ed. 39.94±10.59 n=123 vs Univ 21.00±5.24 n=8 | Also middle 34.80±9.41 n=34, high 34.67±11.25 n=27",
    },
    # Family: Simple vs Extended
    (20, "V6"): {
        "proxy": "Ethnobotanical knowledge score (family type)",
        "n_T": 42, "m_T": 39.81, "sd_T": 11.21,   # Extended family
        "n_C": 158, "m_C": 36.13, "sd_C": 11.28,   # Simple family
        "notas_quant": "t=-1.88 ns | Extended 39.81±11.21 n=42 vs Simple 36.13±11.28 n=158",
    },
    (20, "V1"): {
        "proxy": "Plant species recognized (seed custodianship)",
        "notas_quant": "Sibanda 2025 | Zimbabwe Hwange | 200 informants total | RFC, UF indices | Seed custodian role",
    },

    # ──── ID=45 Ghanimi (2022) Morocco wild edible plants ────
    (45, "V1"): {
        "proxy": "Wild edible plant species richness",
        "notas_quant": "RFC, UF, CF indices | n=130 informants in Messiwa Morocco | 30 table refs",
    },

    # ──── ID=47 Rodriguez-Cruz (2022) Puerto Rico ────
    # Farm production by food security status
    (47, "V4"): {
        "proxy": "Farm production count (food security)",
        "n_T": 124, "m_T": 3.4, "sd_T": 2.5,    # Food secured
        "n_C": 154, "m_C": 3.3, "sd_C": 2.7,     # Persistent food insecure
        "notas_quant": "KW p=0.227 | Secured 3.4±2.5 n=124 vs Persistent 3.3±2.7 n=154 | Also immediate 2.8±2.2 n=123",
    },
    # Network by food security
    (47, "V3"): {
        "proxy": "Social network count (food security)",
        "n_T": 124, "m_T": 2.6, "sd_T": 2.6,    # Food secured
        "n_C": 154, "m_C": 2.0, "sd_C": 1.8,     # Persistent food insecure
        "notas_quant": "KW p=0.140 | Secured 2.6±2.6 n=124 vs Persistent 2.0±1.8 n=154 | Also immediate 2.5±2.0 n=123",
    },
    # Farm size by food security
    (47, "V1"): {
        "proxy": "Farm size (cuerdas, food security)",
        "n_T": 124, "m_T": 75.6, "sd_T": 118.5,  # Food secured (larger farms)
        "n_C": 154, "m_C": 45.1, "sd_C": 81.6,    # Persistent food insecure
        "notas_quant": "KW p=0.029* | Secured 75.6±118.5 vs Persistent 45.1±81.6 | Also immediate 58.0±95",
    },
    # Age by food security
    (47, "V2"): {
        "proxy": "Farmer age (food security)",
        "n_T": 124, "m_T": 52.5, "sd_T": 13.5,   # Food secured
        "n_C": 154, "m_C": 56.4, "sd_C": 13.0,    # Persistent food insecure
        "notas_quant": "KW p=0.053 | Secured 52.5±13.5 vs Persistent 56.4±13.0 | Also immediate 52.6±13.0 n=123",
    },
    # Distance from hurricane eye
    (47, "V6"): {
        "proxy": "Distance from hurricane eye (km)",
        "n_T": 124, "m_T": 25.3, "sd_T": 15.1,   # Food secured (further away)
        "n_C": 154, "m_C": 18.9, "sd_C": 12.3,    # Persistent (closer to eye)
        "notas_quant": "KW p=0.001*** | Secured 25.3±15.1 vs Persistent 18.9±12.3 | Biophysical variable",
    },
    # Landslide exposure
    (47, "V5"): {
        "proxy": "Landslide exposure (km²)",
        "n_T": 124, "m_T": 9.1, "sd_T": 11.6,
        "n_C": 154, "m_C": 12.2, "sd_C": 16.4,
        "notas_quant": "KW p=0.208 | Secured 9.1±11.6 vs Persistent 12.2±16.4 | Biophysical variable",
    },

    # ──── ID=48 Suwardi (Rodriguez 2018 Bolivia ahipa) ────
    (48, "V1"): {
        "proxy": "Ahipa cultivar diversity (MAD)",
        "notas_quant": "Temporal comparison 1994/96 vs 2012 | HB index 0.68-1.63 | p<0.001 ANOVA for density by site | n varies n=3 to n=17 by community",
    },

    # ──── ID=37 Arce (seed networks) ────
    (37, "V1"): {
        "proxy": "Seed transaction volume",
        "notas_quant": "With stress: provisions n=187, acquisitions n=568 | Without stress: provisions n=582, acquisitions n=357 | 17 table refs | Network analysis",
    },
    (37, "V3"): {
        "proxy": "Seed exchange network density",
        "notas_quant": "Potato seed networks | Floury vs bitter landraces | Sales 65-92% of transactions | Multiple comparison groups",
    },

    # ──── ID=32 Ghanimi/Edo (homegarden) ────
    (32, "V1"): {
        "proxy": "Species richness (homegarden)",
        "notas_quant": "n=130 informants | Agroforestry valuation | Shannon/Simpson possible | 30 table refs",
    },

    # ──── ID=43 La Rosa / Aswani (ecological knowledge) ────
    (43, "V1"): {
        "proxy": "Taxonomic distinctness index (ITD)",
        "notas_quant": "Ecological knowledge quantification | Indicator: ITD index | Solomon Islands | Marine and terrestrial taxa",
    },

    # ──── ID=3 de Sousa (biodiversity hotspot) ────
    (3, "V1"): {
        "proxy": "Plant species richness",
        "notas_quant": "p-values: 0.021, 0.259, 0.165, 0.366, 0.002 | 10 tables | Biodiversity hotspot assessment",
    },

    # ──── ID=27 La Rosa/Ferrante (Aegadian Islands) ────
    (27, "V1"): {
        "proxy": "Species richness / ethnobotanical indices",
        "notas_quant": "RFC, CI frequency indices | Aegadian Islands Mediterranean | 13 table refs",
    },
}

# ── Dimension-specific proxy suggestions for studies WITHOUT PDF data ──
DEFAULT_PROXY = {
    "V1": "Agrobiodiversity index / species richness",
    "V2": "Intergenerational knowledge transfer indicator",
    "V3": "Cultural complexity / social network indicator",
    "V4": "Food security / dietary diversity indicator",
    "V5": "Documentation status / knowledge formalization",
    "V6": "External pressure / land use change indicator",
}


def get_abstract_for_study(wb_sel42, study_doi):
    """Get abstract from selecionados_42 by DOI match."""
    ws = wb_sel42.active
    for r in range(2, ws.max_row + 1):
        doi = str(ws.cell(r, 1).value or "")
        if doi and study_doi and doi.lower().strip() == study_doi.lower().strip():
            return str(ws.cell(r, 8).value or "")[:500]
    return ""


def get_pdf_text_first_pages(bdid, pdf_files):
    """Get text from first 3 pages of matched PDF."""
    for fname in pdf_files:
        if fname in EXCLUDE_PDFS:
            continue
        for prefix, bid in PDF_TO_BDID.items():
            if fname.startswith(prefix) and bid == bdid:
                fpath = os.path.join(PDF_DIR, fname)
                try:
                    doc = fitz.open(fpath)
                    text = "".join(page.get_text() for page in doc[:3])
                    doc.close()
                    return text[:3000]
                except:
                    pass
    return ""


def main():
    print("=" * 70)
    print("PREENCHIMENTO COMPLETO V2 (AGRESSIVO)")
    print("=" * 70)

    # Load workbooks
    wb = openpyxl.load_workbook(BD_PREENCHIDO)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    COL = {h: i + 1 for i, h in enumerate(headers)}

    # Load selecionados_42
    wb_sel42 = openpyxl.load_workbook(SEL42_PATH)

    # Get PDF file list
    try:
        pdf_files = sorted([f for f in os.listdir(PDF_DIR) if f.lower().endswith(".pdf")])
    except:
        pdf_files = []
    print(f"PDFs disponiveis: {len(pdf_files)}")

    # Track changes
    changes = {"regiao": 0, "tipo_com": 0, "tipo_int": 0, "proxy": 0,
               "n_T": 0, "notas": 0, "default_proxy": 0}

    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, COL["Study_ID"]).value
        dim = ws.cell(r, COL["Dimensao"]).value
        doi = ws.cell(r, COL["DOI"]).value
        if not sid or not dim:
            continue

        # ── 1. Fill Regiao ──
        if not ws.cell(r, COL["Regiao"]).value and sid in STUDY_META:
            ws.cell(r, COL["Regiao"]).value = STUDY_META[sid]["regiao"]
            ws.cell(r, COL["Regiao"]).fill = HIGHLIGHT_GREEN
            changes["regiao"] += 1

        # ── 2. Fill Tipo_Comunidade ──
        if not ws.cell(r, COL["Tipo_Comunidade"]).value and sid in STUDY_META:
            ws.cell(r, COL["Tipo_Comunidade"]).value = STUDY_META[sid]["tipo_com"]
            ws.cell(r, COL["Tipo_Comunidade"]).fill = HIGHLIGHT_GREEN
            changes["tipo_com"] += 1

        # ── 3. Fill Tipo_Intervencao ──
        if not ws.cell(r, COL["Tipo_Intervencao"]).value and sid in STUDY_META:
            ws.cell(r, COL["Tipo_Intervencao"]).value = STUDY_META[sid]["tipo_int"]
            ws.cell(r, COL["Tipo_Intervencao"]).fill = HIGHLIGHT_GREEN
            changes["tipo_int"] += 1

        # ── 4. Fill quantitative data ──
        key = (sid, dim)
        if key in QUANT_DATA:
            qd = QUANT_DATA[key]

            # Proxy
            if "proxy" in qd and not ws.cell(r, COL["Proxy"]).value:
                ws.cell(r, COL["Proxy"]).value = qd["proxy"]
                ws.cell(r, COL["Proxy"]).fill = HIGHLIGHT
                changes["proxy"] += 1

            # n_T/m_T/sd_T
            if "n_T" in qd and not ws.cell(r, COL["n_T"]).value:
                ws.cell(r, COL["n_T"]).value = qd["n_T"]
                ws.cell(r, COL["m_T"]).value = qd["m_T"]
                ws.cell(r, COL["sd_T"]).value = qd["sd_T"]
                ws.cell(r, COL["n_T"]).fill = HIGHLIGHT
                ws.cell(r, COL["m_T"]).fill = HIGHLIGHT
                ws.cell(r, COL["sd_T"]).fill = HIGHLIGHT
                changes["n_T"] += 1

            # n_C/m_C/sd_C
            if "n_C" in qd and qd.get("n_C") and not ws.cell(r, COL["n_C"]).value:
                ws.cell(r, COL["n_C"]).value = qd["n_C"]
                ws.cell(r, COL["m_C"]).value = qd["m_C"]
                ws.cell(r, COL["sd_C"]).value = qd["sd_C"]
                ws.cell(r, COL["n_C"]).fill = HIGHLIGHT
                ws.cell(r, COL["m_C"]).fill = HIGHLIGHT
                ws.cell(r, COL["sd_C"]).fill = HIGHLIGHT

            # Update Notas with quantitative detail
            if "notas_quant" in qd:
                existing = ws.cell(r, COL["Notas"]).value or ""
                new_nota = qd["notas_quant"]
                if new_nota not in existing:
                    combined = f"[QUANT] {new_nota} || {existing}" if existing else f"[QUANT] {new_nota}"
                    ws.cell(r, COL["Notas"]).value = combined
                    ws.cell(r, COL["Notas"]).fill = HIGHLIGHT
                    changes["notas"] += 1

        # ── 5. Default Proxy if still empty ──
        if not ws.cell(r, COL["Proxy"]).value and dim in DEFAULT_PROXY:
            ws.cell(r, COL["Proxy"]).value = DEFAULT_PROXY[dim]
            ws.cell(r, COL["Proxy"]).fill = HIGHLIGHT_GREEN
            changes["default_proxy"] += 1

        # ── 6. Fill Notas with abstract for studies without PDF ──
        if not ws.cell(r, COL["Notas"]).value and doi:
            abstract = get_abstract_for_study(wb_sel42, doi)
            if abstract:
                ws.cell(r, COL["Notas"]).value = f"[ABSTRACT] {abstract}"
                ws.cell(r, COL["Notas"]).fill = HIGHLIGHT_GREEN
                changes["notas"] += 1

    wb_sel42.close()

    # Save
    wb.save(OUT_PATH)
    print(f"\nSALVO: {OUT_PATH}")

    # Verification
    print(f"\n{'='*70}")
    print("MUDANÇAS REALIZADAS:")
    for k, v in changes.items():
        print(f"  {k:<20}: {v}")

    # Final column fill report
    wb2 = openpyxl.load_workbook(OUT_PATH)
    ws2 = wb2.active
    print(f"\n{'='*70}")
    print("PREENCHIMENTO FINAL:")
    total = ws2.max_row - 1
    for col_name in ["Proxy", "n_T", "m_T", "sd_T", "n_C", "m_C", "sd_C",
                     "Tipo_Intervencao", "Regiao", "Tipo_Comunidade", "Notas",
                     "Tempo_Intervencao", "NOS"]:
        ci = COL[col_name]
        filled = sum(1 for r in range(2, ws2.max_row + 1)
                     if ws2.cell(r, ci).value is not None and str(ws2.cell(r, ci).value).strip())
        print(f"  {col_name:<25}: {filled:3d}/{total} ({filled/total*100:.1f}%)")
    wb2.close()


if __name__ == "__main__":
    main()
