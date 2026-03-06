#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
30_preparar_codificacao_quali.py
================================
Lê bd_extracao_PREENCHIDO.xlsx, classifica cada registro por tier de evidência
e gera bd_codificacao_qualitativa.xlsx com colunas para codificação manual
(dupla-cega) de direção e intensidade do efeito.

Tiers:
  T1  — Quantitativo completo (n_T, m_T, sd_T, n_C, m_C, sd_C preenchidos)
  T2a — p-values numéricos extraídos + n disponível (parcial ou completo)
  T2b — [TEM_STAT] ANOVA/KW marcado, sem p-values explícitos
  T3  — Results=SIM com comparações listadas, sem estatísticas
  T4  — [QUALIT] puro
  EX  — [SEM ACESSO] ou Results=NAO  → excluído

Saída: bd_codificacao_qualitativa.xlsx  (para preenchimento manual)
       diagnostico_tiers.xlsx           (resumo por tier × dimensão)

Autor: Diego Vidal  |  2026-03-02
"""

import re
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ── Caminhos ──────────────────────────────────────────────────────
BASE = os.path.dirname(os.path.abspath(__file__))
DIR_DADOS = os.path.join(BASE, "..", "2-BANCO_DADOS", "2-DADOS_TABULADOS")
FP_IN  = os.path.join(DIR_DADOS, "bd_extracao_PREENCHIDO.xlsx")
FP_OUT = os.path.join(DIR_DADOS, "bd_codificacao_qualitativa.xlsx")
FP_DIAG = os.path.join(DIR_DADOS, "diagnostico_tiers.xlsx")


def extrair_pvalues(notas: str):
    """Extrai p-values numéricos de strings como "p-values: ['0.021', '0.002']" """
    pvals = []
    # Padrão: lista Python-like
    m = re.search(r"p-values:\s*\[([^\]]+)\]", notas)
    if m:
        for token in re.findall(r"[\d.]+", m.group(1)):
            try:
                pvals.append(float(token))
            except ValueError:
                pass
    # Padrão: p < 0.05, p = 0.001, etc.
    for m2 in re.finditer(r"p\s*[<=]\s*([\d.]+)", notas):
        try:
            pvals.append(float(m2.group(1)))
        except ValueError:
            pass
    return list(set(pvals))  # deduplica


def extrair_n_from_notas(notas: str):
    """Extrai n reportado das notas (ex: 'n=[130]' ou 'n=130 informants')"""
    ns = []
    # Padrão array: n=[130, 200]
    m = re.search(r"n=\[([^\]]+)\]", notas)
    if m:
        for token in re.findall(r"\d+", m.group(1)):
            ns.append(int(token))
    # Padrão simples: n=130
    for m2 in re.finditer(r"\bn=(\d+)\b", notas):
        ns.append(int(m2.group(1)))
    # sample sizes: [68, 151, ...]
    m3 = re.search(r"sample sizes:\s*\[([^\]]+)\]", notas)
    if m3:
        for token in re.findall(r"\d+", m3.group(1)):
            ns.append(int(token))
    return list(set(ns))


def classificar_tier(row):
    """Classifica o tier de um registro baseado nos campos disponíveis."""
    n_T, m_T, sd_T = row[9], row[10], row[11]
    n_C, m_C, sd_C = row[12], row[13], row[14]
    notas = str(row[20]) if row[20] else ""

    # T1: tudo preenchido
    if all(v is not None for v in [n_T, m_T, sd_T, n_C, m_C, sd_C]):
        return "T1"

    # Excluídos
    if "[SEM ACESSO]" in notas:
        return "EX_SEM_ACESSO"
    if "Results=NAO" in notas and "p-values" not in notas and "[TEM_STAT]" not in notas:
        return "EX_SEM_RESULTADO"

    # T2a: tem p-values extraídos
    pvals = extrair_pvalues(notas)
    if pvals:
        return "T2a"

    # T2b: TEM_STAT ou ANOVA/KW
    if "[TEM_STAT]" in notas or "ANOVA" in notas or "KW" in notas or "Kruskal" in notas:
        return "T2b"

    # T3: Results=SIM com comparações
    if "Results=SIM" in notas:
        return "T3"

    # T4: QUALIT puro
    if "[QUALIT]" in notas or "[QUAL]" in notas:
        return "T4"

    # Outros com alguma info
    if notas.strip():
        return "T3"  # conservador: assume que há algo para codificar

    return "EX_VAZIO"


def extrair_comparacoes(notas: str):
    """Extrai tipos de comparação listados."""
    m = re.search(r"comparisons:\s*(.+?)(?:\||$)", notas)
    if m:
        return m.group(1).strip()
    return ""


def main():
    wb_in = openpyxl.load_workbook(FP_IN, read_only=True)
    ws_in = wb_in["DADOS"]

    # ── Ler todos os registros ──
    registros = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        notas = str(row[20]) if row[20] else ""
        tier = classificar_tier(row)
        pvals = extrair_pvalues(notas) if tier == "T2a" else []
        ns = extrair_n_from_notas(notas)
        comps = extrair_comparacoes(notas)

        registros.append({
            "Study_ID": row[0],
            "Study": row[1],
            "DOI": row[2],
            "Year": row[5],
            "Dimensao": row[6],
            "Dimensao_Label": row[7],
            "Proxy": row[8],
            "n_T": row[9],
            "m_T": row[10],
            "sd_T": row[11],
            "n_C": row[12],
            "m_C": row[13],
            "sd_C": row[14],
            "Tipo_Intervencao": row[15],
            "Regiao": row[16],
            "Tipo_Comunidade": row[17],
            "NOS": row[19],
            "Notas": notas[:800],
            "Tier": tier,
            "p_values_extraidos": "; ".join(f"{p:.6f}" for p in pvals) if pvals else "",
            "n_reportados": "; ".join(str(n) for n in ns) if ns else "",
            "Comparacoes_extraidas": comps,
            # ── Colunas para codificação manual (PREENCHER) ──
            "Direcao_efeito": "",  # +1 (vulnerabilidade agravada), 0 (neutro), -1 (redução)
            "Intensidade": "",     # 1 (fraca), 2 (moderada), 3 (forte)
            "n_T_codificado": "",  # n do grupo/condição tratamento (se extraível do PDF)
            "n_C_codificado": "",  # n do grupo/condição controle
            "Revisor": "",         # R1 ou R2
            "Confianca_codificacao": "",  # Alta, Moderada, Baixa
            "Notas_codificador": "",
        })

    wb_in.close()

    # ── Diagnóstico por Tier × Dimensão ──
    diag = defaultdict(lambda: defaultdict(int))
    for r in registros:
        diag[r["Tier"]][r["Dimensao"]] += 1
        diag[r["Tier"]]["TOTAL"] += 1

    # ── Gerar planilha de codificação ──
    wb_out = openpyxl.Workbook()

    # Sheet 1: Registros para codificação (excluindo T1 e EX)
    ws_cod = wb_out.active
    ws_cod.title = "CODIFICACAO"

    headers = [
        "Study_ID", "Study", "DOI", "Year", "Dimensao", "Dimensao_Label",
        "Proxy", "Tier",
        "p_values_extraidos", "n_reportados", "Comparacoes_extraidas",
        "Notas",
        "Direcao_efeito", "Intensidade",
        "n_T_codificado", "n_C_codificado",
        "Revisor", "Confianca_codificacao", "Notas_codificador",
    ]

    # Estilos
    fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    fill_manual = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    font_header = Font(bold=True, color="FFFFFF", size=10)
    font_body = Font(size=9)

    # Cabeçalho
    for j, h in enumerate(headers, 1):
        cell = ws_cod.cell(row=1, column=j, value=h)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Dados (apenas tiers que precisam de codificação: T2a, T2b, T3, T4)
    row_i = 2
    codificaveis = [r for r in registros if r["Tier"] in ("T2a", "T2b", "T3", "T4")]
    for r in codificaveis:
        vals = [r.get(h, "") for h in headers]
        for j, v in enumerate(vals, 1):
            cell = ws_cod.cell(row=row_i, column=j, value=v)
            cell.font = font_body
            # Destacar colunas manuais
            if headers[j-1] in ("Direcao_efeito", "Intensidade", "n_T_codificado",
                                "n_C_codificado", "Revisor", "Confianca_codificacao",
                                "Notas_codificador"):
                cell.fill = fill_manual
        row_i += 1

    # Ajustar larguras
    col_widths = {
        "A": 8, "B": 18, "C": 25, "D": 6, "E": 8, "F": 22,
        "G": 30, "H": 6, "I": 20, "J": 15, "K": 30,
        "L": 60, "M": 14, "N": 10, "O": 12, "P": 12,
        "Q": 8, "R": 18, "S": 30,
    }
    for col, w in col_widths.items():
        ws_cod.column_dimensions[col].width = w

    # Sheet 2: Registros T1 (quantitativos, referência)
    ws_t1 = wb_out.create_sheet("T1_QUANTITATIVOS")
    headers_t1 = [
        "Study_ID", "Study", "Dimensao", "Proxy",
        "n_T", "m_T", "sd_T", "n_C", "m_C", "sd_C",
        "Tipo_Intervencao", "Regiao", "Tipo_Comunidade",
    ]
    for j, h in enumerate(headers_t1, 1):
        cell = ws_t1.cell(row=1, column=j, value=h)
        cell.fill = fill_header
        cell.font = font_header

    row_i = 2
    for r in registros:
        if r["Tier"] != "T1":
            continue
        for j, h in enumerate(headers_t1, 1):
            ws_t1.cell(row=row_i, column=j, value=r.get(h, "")).font = font_body
        row_i += 1

    # Sheet 3: Excluídos
    ws_ex = wb_out.create_sheet("EXCLUIDOS")
    headers_ex = ["Study_ID", "Study", "Dimensao", "Proxy", "Tier", "Notas"]
    for j, h in enumerate(headers_ex, 1):
        cell = ws_ex.cell(row=1, column=j, value=h)
        cell.fill = fill_header
        cell.font = font_header

    row_i = 2
    for r in registros:
        if not r["Tier"].startswith("EX"):
            continue
        for j, h in enumerate(headers_ex, 1):
            ws_ex.cell(row=row_i, column=j, value=r.get(h, "")).font = font_body
        row_i += 1

    # Sheet 4: Instruções de codificação
    ws_instr = wb_out.create_sheet("INSTRUCOES")
    instrucoes = [
        ("Campo", "Descrição"),
        ("Direcao_efeito",
         "+1 = vulnerabilidade agravada (indicador piorou no grupo/período tratamento); "
         "0 = neutro/sem diferença reportada; "
         "-1 = vulnerabilidade reduzida (indicador melhorou). "
         "ATENÇÃO: a direção refere-se à VULNERABILIDADE, não ao indicador bruto. "
         "Ex: aumento de diversidade = redução de vulnerabilidade = -1."),
        ("Intensidade",
         "1 = efeito fraco (diferença marginal, não significativa ou p > 0.10); "
         "2 = efeito moderado (diferença significativa p < 0.05 ou moderada sem teste); "
         "3 = efeito forte (grande diferença, p < 0.01, ou mudança qualitativa clara)."),
        ("n_T_codificado",
         "Tamanho amostral do grupo tratamento/intervenção extraído do PDF. "
         "Se o estudo reporta n total sem separar grupos, colocar n/2 e anotar."),
        ("n_C_codificado",
         "Tamanho amostral do grupo controle/comparação. Mesmo critério acima."),
        ("Confianca_codificacao",
         "Alta = dados claros no texto/tabela; Moderada = inferido de figuras/contexto; "
         "Baixa = interpretação subjetiva de narrativa qualitativa."),
        ("", ""),
        ("PROTOCOLO", ""),
        ("1",
         "Ler abstract + resultados + tabelas do estudo original."),
        ("2",
         "Identificar a comparação mais relevante para a dimensão V1-V6 atribuída."),
        ("3",
         "Codificar Direcao e Intensidade de forma independente (R1 e R2)."),
        ("4",
         "Calcular κ de Cohen após codificação completa. Resolver discordâncias por consenso."),
        ("5",
         "Para Tier T2a com p-values: a direção determina o sinal do efeito convertido."),
        ("", ""),
        ("MAPEAMENTO ORDINAL → lnOR",
         "Após codificação, o script 31 converte: "
         "Intensidade 1 → quantil 0.40 da logit-normal (lnOR = -0.405); "
         "Intensidade 2 → quantil 0.60 (lnOR = +0.405); "
         "Intensidade 3 → quantil 0.80 (lnOR = +1.386). "
         "O sinal é multiplicado por Direcao_efeito. "
         "Variância fixa = π²/3 ≈ 3.290 para todos os ordinais (Hasselblad & Hedges 1995)."),
    ]
    for i, (campo, desc) in enumerate(instrucoes, 1):
        ws_instr.cell(row=i, column=1, value=campo).font = Font(bold=True, size=10)
        ws_instr.cell(row=i, column=2, value=desc).font = Font(size=9)
    ws_instr.column_dimensions["A"].width = 25
    ws_instr.column_dimensions["B"].width = 80

    # Sheet 5: Diagnóstico
    ws_diag = wb_out.create_sheet("DIAGNOSTICO")
    dims = ["V1", "V2", "V3", "V4", "V5", "V6", "TOTAL"]
    tiers_order = ["T1", "T2a", "T2b", "T3", "T4", "EX_SEM_ACESSO", "EX_SEM_RESULTADO", "EX_VAZIO"]

    ws_diag.cell(row=1, column=1, value="Tier").font = font_header
    for j, d in enumerate(dims, 2):
        ws_diag.cell(row=1, column=j, value=d).font = font_header

    for i, tier in enumerate(tiers_order, 2):
        ws_diag.cell(row=i, column=1, value=tier).font = Font(bold=True)
        for j, d in enumerate(dims, 2):
            ws_diag.cell(row=i, column=j, value=diag.get(tier, {}).get(d, 0))

    wb_out.save(FP_OUT)
    print(f"✔ Planilha de codificação salva em: {FP_OUT}")
    print(f"  → {len(codificaveis)} registros para codificação manual")
    print(f"  → {sum(1 for r in registros if r['Tier'] == 'T1')} registros T1 (quantitativos)")
    print(f"  → {sum(1 for r in registros if r['Tier'].startswith('EX'))} registros excluídos")

    # ── Diagnóstico separado ──
    wb_diag = openpyxl.Workbook()
    ws = wb_diag.active
    ws.title = "Tiers"
    ws.cell(row=1, column=1, value="Tier")
    for j, d in enumerate(dims, 2):
        ws.cell(row=1, column=j, value=d)
    for i, tier in enumerate(tiers_order, 2):
        ws.cell(row=i, column=1, value=tier)
        for j, d in enumerate(dims, 2):
            ws.cell(row=i, column=j, value=diag.get(tier, {}).get(d, 0))
    wb_diag.save(FP_DIAG)
    print(f"✔ Diagnóstico salvo em: {FP_DIAG}")

    # ── Resumo console ──
    print("\n=== DIAGNÓSTICO POR TIER ===")
    for tier in tiers_order:
        total = diag.get(tier, {}).get("TOTAL", 0)
        by_dim = " | ".join(f"{d}={diag.get(tier, {}).get(d, 0)}" for d in dims[:6])
        print(f"  {tier:20s}  total={total:3d}  ({by_dim})")


if __name__ == "__main__":
    main()
