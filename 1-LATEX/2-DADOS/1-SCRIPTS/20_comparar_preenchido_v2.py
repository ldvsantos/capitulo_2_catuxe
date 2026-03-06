# -*- coding: utf-8 -*-
"""
20_comparar_preenchido_v2.py
============================
Compara bd_extracao_PREENCHIDO.xlsx com bd_extracao_V2.xlsx
para identificar lacunas a preencher.
"""
import openpyxl, shutil, tempfile, sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

TAB = Path(r"c:\Users\vidal\OneDrive\Documentos\13 - CLONEGIT"
           r"\artigo_2_catuxe\1-LATEX\2-DADOS\2-BANCO_DADOS\2-DADOS_TABULADOS")

# Copiar para temp
tmp_p = Path(tempfile.gettempdir()) / "pre_copy.xlsx"
tmp_v = Path(tempfile.gettempdir()) / "v2_copy.xlsx"
shutil.copy2(TAB / "bd_extracao_PREENCHIDO.xlsx", tmp_p)
shutil.copy2(TAB / "bd_extracao_V2.xlsx", tmp_v)

# Ler PREENCHIDO
wb_p = openpyxl.load_workbook(tmp_p, read_only=True)
ws_p = wb_p.active
hdr_p = [c.value for c in next(ws_p.iter_rows(min_row=1, max_row=1))]
rows_p = [list(r) for r in ws_p.iter_rows(min_row=2, values_only=True)]
wb_p.close()

# Ler V2
wb_v = openpyxl.load_workbook(tmp_v, read_only=True)
ws_v = wb_v.active
hdr_v = [c.value for c in next(ws_v.iter_rows(min_row=1, max_row=1))]
rows_v = [list(r) for r in ws_v.iter_rows(min_row=2, values_only=True)]
wb_v.close()

print("PREENCHIDO header:", hdr_p)
print(f"PREENCHIDO rows: {len(rows_p)}")
print()
print("V2 header:", hdr_v)
print(f"V2 rows: {len(rows_v)}")
print()

# Comparar colunas
print("=" * 100)
print("COBERTURA POR COLUNA")
print("=" * 100)

def count_filled(rows, col_idx):
    n = 0
    for r in rows:
        if col_idx < len(r) and r[col_idx] is not None and str(r[col_idx]).strip() != "":
            n += 1
    return n

for i, col_name in enumerate(hdr_p):
    fp = count_filled(rows_p, i)
    # Find same column in V2
    if col_name in hdr_v:
        j = hdr_v.index(col_name)
        fv = count_filled(rows_v, j)
    else:
        fv = -1
    total_p = len(rows_p)
    total_v = len(rows_v)
    delta = fv - fp if fv >= 0 else 0
    marker = " <<<" if delta > 0 else ""
    print(f"  {col_name:<20}  PREENCHIDO: {fp:3d}/{total_p}  V2: {fv:3d}/{total_v}  delta: +{delta}{marker}")

# Colunas extras em V2 ausentes em PREENCHIDO
extras_v2 = [c for c in hdr_v if c not in hdr_p]
if extras_v2:
    print(f"\nColunas em V2 ausentes em PREENCHIDO: {extras_v2}")
    for col_name in extras_v2:
        j = hdr_v.index(col_name)
        fv = count_filled(rows_v, j)
        print(f"  {col_name:<20}  V2: {fv:3d}/{len(rows_v)}")

# Verificar estudos em V2 que não estão em PREENCHIDO
ids_p = set()
for r in rows_p:
    if r[0] is not None:
        ids_p.add(r[0])

ids_v = set()
for r in rows_v:
    if r[0] is not None:
        ids_v.add(r[0])

missing_ids = ids_v - ids_p
extra_ids = ids_p - ids_v
print(f"\nIDs em V2 ausentes em PREENCHIDO: {sorted(missing_ids)}")
print(f"IDs em PREENCHIDO ausentes em V2: {sorted(extra_ids)}")

# Detalhar células vazias em PREENCHIDO que têm dados em V2
print("\n" + "=" * 100)
print("CÉLULAS A ATUALIZAR (vazias em PREENCHIDO, preenchidas em V2)")
print("=" * 100)

# Build V2 lookup: (study_id, dimensao) -> row
v2_lookup = {}
dim_col_v = hdr_v.index("Dimensao") if "Dimensao" in hdr_v else 6
for r in rows_v:
    key = (r[0], r[dim_col_v])
    v2_lookup[key] = r

dim_col_p = hdr_p.index("Dimensao") if "Dimensao" in hdr_p else 6

update_count = 0
for r_p in rows_p:
    sid = r_p[0]
    dim = r_p[dim_col_p]
    key = (sid, dim)
    r_v = v2_lookup.get(key)
    if r_v is None:
        continue
    for i, col_name in enumerate(hdr_p):
        if col_name in hdr_v:
            j = hdr_v.index(col_name)
            val_p = r_p[i]
            val_v = r_v[j] if j < len(r_v) else None
            is_empty_p = val_p is None or str(val_p).strip() == ""
            is_filled_v = val_v is not None and str(val_v).strip() != ""
            if is_empty_p and is_filled_v:
                update_count += 1

print(f"Total células a atualizar: {update_count}")
