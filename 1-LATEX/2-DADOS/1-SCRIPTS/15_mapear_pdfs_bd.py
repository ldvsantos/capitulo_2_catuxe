# -*- coding: utf-8 -*-
"""
15_mapear_pdfs_bd.py
====================
Mapeia os 27 PDFs aos Study_IDs corretos do bd_extracao.xlsx
usando DOIs extraídos das primeiras páginas de cada PDF.
"""

import os, re, json
from pathlib import Path

try:
    import fitz
except ImportError:
    import sys; sys.exit("pip install PyMuPDF")

import openpyxl

# Caminhos
BASE = Path(r"c:\Users\vidal\OneDrive\Documentos\13 - CLONEGIT"
            r"\artigo_2_catuxe\1-LATEX\2-DADOS\2-BANCO_DADOS")
PDF_DIR = r"\\?\ ".strip() + str(BASE / "1-ARTIGOS_SELECIONADOS")
BD_PATH = BASE / "2-DADOS_TABULADOS" / "bd_extracao.xlsx"
OUT_DIR = BASE / "3-OUTPUT"

# 1. Read DOIs from bd_extracao.xlsx
wb = openpyxl.load_workbook(BD_PATH)
ws = wb.active
db_dois = {}
seen = set()
for r in range(2, ws.max_row + 1):
    sid = ws.cell(r, 1).value
    if sid in seen:
        continue
    seen.add(sid)
    doi = str(ws.cell(r, 3).value or "").strip().lower()
    study = str(ws.cell(r, 2).value or "")
    title = str(ws.cell(r, 4).value or "")
    if doi:
        db_dois[doi] = {"id": sid, "study": study, "title": title}

print(f"DOIs no bd_extracao: {len(db_dois)}")

# 2. DOI pattern
doi_pat = re.compile(r"10\.\d{4,9}/[^\s,;\"')>\]]+")

# 3. Process PDFs
pdf_matches = []
for fname in sorted(os.listdir(PDF_DIR)):
    if not fname.lower().endswith(".pdf"):
        continue
    fpath = os.path.join(PDF_DIR, fname)
    doc = fitz.open(fpath)
    text = ""
    for i, page in enumerate(doc):
        if i >= 3:
            break
        text += page.get_text()
    doc.close()

    found_dois = doi_pat.findall(text)
    clean_dois = set()
    for d in found_dois:
        d = d.rstrip(".,;:")
        clean_dois.add(d.lower())

    matched_id = None
    matched_doi = None
    matched_study = None
    for d in clean_dois:
        if d in db_dois:
            matched_id = db_dois[d]["id"]
            matched_doi = d
            matched_study = db_dois[d]["study"]
            break
    if not matched_id:
        for d in clean_dois:
            for db_d, info in db_dois.items():
                if db_d in d or d in db_d:
                    matched_id = info["id"]
                    matched_doi = db_d
                    matched_study = info["study"]
                    break
            if matched_id:
                break

    status = (f"MATCH ID={matched_id:2d} {matched_study[:30]}"
              if matched_id else "NO MATCH")
    pdf_matches.append({
        "file": fname,
        "matched_id": matched_id,
        "matched_doi": matched_doi,
        "sample_dois": sorted(list(clean_dois))[:5],
    })
    print(f"  {fname[:57]:<59} -> {status}")

matched = [p for p in pdf_matches if p["matched_id"]]
unmatched = [p for p in pdf_matches if not p["matched_id"]]
print(f"\nMatched: {len(matched)}/27")
print(f"Unmatched: {len(unmatched)}")

if unmatched:
    print("\nUNMATCHED PDFs (sample DOIs found):")
    for u in unmatched:
        print(f"  {u['file'][:70]}")
        for d in u["sample_dois"][:5]:
            print(f"    DOI: {d}")

# 4. Also find which bd_extracao studies have NO pdf
matched_ids = {p["matched_id"] for p in matched}
print(f"\nStudy_IDs COM PDF: {sorted(matched_ids)}")
no_pdf = sorted(set(range(1, 43)) - matched_ids)
print(f"Study_IDs SEM PDF ({len(no_pdf)}): {no_pdf}")

# 5. Save mapping
mapping = {p["file"]: p["matched_id"] for p in matched}
map_path = OUT_DIR / "pdf_to_bdid_map.json"
with open(map_path, "w") as f:
    json.dump(
        {"matched": pdf_matches, "unmatched_ids": no_pdf},
        f, indent=2, ensure_ascii=False
    )
print(f"\nMapeamento salvo: {map_path}")
