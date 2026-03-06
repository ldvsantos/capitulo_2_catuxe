# -*- coding: utf-8 -*-
"""25_validacao_final.py"""
import sys
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

TAB = Path(r"c:\Users\vidal\OneDrive\Documentos\13 - CLONEGIT"
           r"\artigo_2_catuxe\1-LATEX\2-DADOS\2-BANCO_DADOS\2-DADOS_TABULADOS")
BD_PATH = TAB / "bd_extracao_PREENCHIDO.xlsx"
OUT = TAB / "validacao_final.txt"

L = []
def p(s=""):
    L.append(str(s))

wb = openpyxl.load_workbook(BD_PATH)
ws = wb.active
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
COL = {h: i + 1 for i, h in enumerate(headers)}
total = ws.max_row - 1

p("=" * 70)
p("VALIDACAO FINAL - bd_extracao_PREENCHIDO.xlsx")
p("=" * 70)

p("\n-- COBERTURA POR COLUNA --")
for cn in ["Study_ID","Dimensao","Proxy","n_T","m_T","sd_T","n_C","m_C","sd_C","Tipo_Intervencao","Regiao","Tipo_Comunidade","Notas"]:
    ci = COL.get(cn)
    if not ci:
        p(f"  {cn}: NAO ENCONTRADA"); continue
    filled = sum(1 for r in range(2, ws.max_row+1) if ws.cell(r,ci).value is not None and str(ws.cell(r,ci).value).strip())
    p(f"  {cn:<25}: {filled:3d}/{total} ({filled/total*100:.1f}%)")

p("\n-- ESTUDOS COM MEAN+/-SD --")
ids_quant = {}
for r in range(2, ws.max_row+1):
    sid = ws.cell(r, COL["Study_ID"]).value
    dim = ws.cell(r, COL["Dimensao"]).value
    n_t = ws.cell(r, COL["n_T"]).value
    if n_t is not None and str(n_t).strip():
        ids_quant.setdefault(sid, []).append({
            "dim": dim,
            "n_T": n_t, "m_T": ws.cell(r,COL["m_T"]).value,
            "sd_T": ws.cell(r,COL["sd_T"]).value,
            "n_C": ws.cell(r,COL["n_C"]).value,
            "m_C": ws.cell(r,COL["m_C"]).value,
            "sd_C": ws.cell(r,COL["sd_C"]).value,
            "proxy": str(ws.cell(r,COL["Proxy"]).value or "")[:60],
        })

for sid in sorted(ids_quant):
    entries = ids_quant[sid]
    p(f"\n  ID={sid} ({len(entries)} dim):")
    for e in entries:
        sd_t = float(e['sd_T']) if e['sd_T'] else 0
        sd_c = float(e['sd_C']) if e['sd_C'] else 0
        p(f"    {e['dim']}: T={e['m_T']}+/-{sd_t:.2f} (n={e['n_T']}) vs C={e['m_C']}+/-{sd_c:.2f} (n={e['n_C']}) | {e['proxy']}")

n_quant_rows = sum(len(v) for v in ids_quant.values())
p(f"\n  Total: {len(ids_quant)} estudos, {n_quant_rows} linhas com mean+/-SD")

p("\n-- ESTUDOS SEM ACESSO A PDF --")
ids_sem = set()
for r in range(2, ws.max_row+1):
    notas = str(ws.cell(r, COL["Notas"]).value or "")
    if "[SEM ACESSO]" in notas:
        ids_sem.add(ws.cell(r, COL["Study_ID"]).value)
p(f"  IDs: {sorted(ids_sem)} ({len(ids_sem)} estudos, {len(ids_sem)*6} linhas)")

p("\n-- TAGS NAS NOTAS --")
tag_map = {}
for r in range(2, ws.max_row+1):
    sid = ws.cell(r, COL["Study_ID"]).value
    notas = str(ws.cell(r, COL["Notas"]).value or "")
    found = False
    for tag in ["[IDEAL]","[QUALIT]","[TEM_STAT]","[TEM_P]","[STAT]","[QUANT]","[ABSTRACT]","[SEM ACESSO]","[QUAL]"]:
        if tag in notas:
            tag_map.setdefault(tag, set()).add(sid)
            found = True
    if not found:
        tag_map.setdefault("SEM_TAG", set()).add(sid)

for tag in sorted(tag_map):
    ids = tag_map[tag]
    p(f"  {tag:<15}: {len(ids)} estudos -> IDs {sorted(ids)}")

p("\n-- RESUMO META-ANALISE --")
n_total = 48
n_sem = len(ids_sem)
n_com = n_total - n_sem
n_ideal = len(ids_quant)
n_qual = n_com - n_ideal
p(f"  Total: {n_total}")
p(f"  Sem acesso PDF: {n_sem} ({sorted(ids_sem)})")
p(f"  Com acesso PDF: {n_com}")
p(f"  Com mean+/-SD: {n_ideal} estudos, {n_quant_rows} linhas")
p(f"  Sem mean+/-SD: {n_qual}")
if n_com > 0:
    p(f"  Cobertura quant: {n_quant_rows}/{n_com*6} ({n_quant_rows/(n_com*6)*100:.1f}%)")

wb.close()
p("\nOK.")

text = "\n".join(L)
OUT.write_text(text, encoding="utf-8")
print(f"Salvo: {OUT}")
