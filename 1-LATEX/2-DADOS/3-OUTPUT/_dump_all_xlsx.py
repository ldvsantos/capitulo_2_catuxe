import openpyxl
import os
import sys

output_dir = os.path.dirname(os.path.abspath(__file__))

files_to_read = [
    "resultados_por_dimensao.xlsx",
    "subgrupos.xlsx",
    "vies_publicacao.xlsx",
    "trimfill.xlsx",
    "sensibilidade_DL_REML.xlsx",
    "sensibilidade_rho.xlsx",
    "sensibilidade_tier.xlsx",
    "leave_one_out.xlsx",
    "ranking_lnRR.xlsx",
    "mapa_heterogeneidade.xlsx",
    "coeficientes_meta_regressao.xlsx",
    "moderadores_exploratoria.xlsx",
    "moderadores_significativos.xlsx",
    "diagnostico_evidencia_mista.xlsx",
    "tabela_ISB_consolidada.xlsx",
    "universos_discurso_fuzzy.xlsx",
]

for fname in files_to_read:
    fpath = os.path.join(output_dir, fname)
    if not os.path.exists(fpath):
        print(f"\n{'='*80}")
        print(f"FILE NOT FOUND: {fname}")
        print(f"{'='*80}")
        continue
    
    print(f"\n{'='*80}")
    print(f"FILE: {fname}")
    print(f"{'='*80}")
    
    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols) ---")
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
                vals = []
                for cell in row:
                    v = cell.value
                    if v is None:
                        vals.append("")
                    else:
                        vals.append(str(v))
                print("\t".join(vals))
        wb.close()
    except Exception as e:
        print(f"ERROR reading {fname}: {e}")

print("\n\nDONE.")
